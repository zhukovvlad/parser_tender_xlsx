# helpers/get_summary.py

"""Модуль для извлечения общих итоговых данных по тендеру.

Назначение:
    Этот модуль предоставляет специализированную функцию `get_summary`,
    которая отвечает за парсинг блока с итоговыми строками (summary)
    в тендерной таблице.

    Ключевой особенностью является то, что этот блок содержит агрегированные
    данные по всему тендеру (например, "ИТОГО С НДС", "В том числе НДС")
    и его расположение не привязано к границам конкретных лотов. Функция
    самостоятельно находит этот блок в нижней части таблицы и извлекает из него
    данные для указанного подрядчика.

Ключевые зависимости:
    - `parse_contractor_row`: для извлечения числовых данных из колонок,
      относящихся к конкретному подрядчику.
"""

import logging
from typing import Any, Dict

from openpyxl.worksheet.worksheet import Worksheet

from ..constants import (
    JSON_KEY_DEVIATION_FROM_CALCULATED_COST,
    JSON_KEY_INITIAL_COST,
    JSON_KEY_JOB_TITLE,
    JSON_KEY_TOTAL_COST_VAT,
    JSON_KEY_VAT,
    START_INDEXING_POSITION_ROW,
    TABLE_PARSE_DEVIATION_FROM_CALCULATED_COST,
    TABLE_PARSE_INITIAL_COST,
)
from .parse_contractor_row import parse_contractor_row

log = logging.getLogger(__name__)


def get_summary(ws: Worksheet, contractor: Dict[str, Any]) -> Dict[str, Any]:
    """Извлекает итоговые (summary) строки для подрядчика со всего листа.

    Функция предназначена для парсинга глобального блока итогов, который,
    как правило, находится внизу основной таблицы с позициями.

    Логика работы:
    1.  **Поиск начальной строки:** Функция итерируется по строкам листа,
        начиная с `START_INDEXING_POSITION_ROW`. Началом блока итогов
        считается первая встреченная строка, в которой ячейка из колонки 'A'
        является частью объединенного диапазона.
    2.  **Обработка блока:** После нахождения начальной строки, функция
        продолжает считывать все последующие строки до тех пор, пока не
        встретит первую полностью пустую строку, которая сигнализирует
        о конце блока итогов.
    3.  **Парсинг строк:** Для каждой строки в блоке:
        - Определяется семантический ключ (`summary_key`) на основе
          ключевых слов в текстовом названии (например, "итого", "ндс").
          Если ключевые слова не найдены, используется общий ключ
          вида "merged_{номер_строки}".
        - Вызывается `parse_contractor_row` для извлечения числовых
          данных из колонок, относящихся к переданному подрядчику.
        - Формируется словарь с данными для этой итоговой строки.

    Args:
        ws (Worksheet): Активный лист Excel для анализа.
        contractor (Dict[str, Any]): Словарь с информацией о подрядчике,
            необходимый для вызова `parse_contractor_row`.

    Returns:
        Dict[str, Any]: Словарь, где ключи идентифицируют итоговую строку
        (например, "total_cost_vat", "vat"), а значения - словари
        с данными для этой строки. Если блок итогов не найден,
        возвращается пустой словарь.
    """
    summary: Dict[str, Any] = {}
    summary_start_row = -1

    # Находим, где начинается блок summary
    for row_num in range(START_INDEXING_POSITION_ROW, ws.max_row + 1):
        first_cell = ws.cell(row=row_num, column=1)

        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if first_cell.coordinate in merged_range:
                is_merged = True
                break
        if is_merged:
            summary_start_row = row_num
            break

    if summary_start_row == -1:
        log.debug("Блок summary не найден на листе.")
        return {}  # Блок итогов не найден

    # Считываем все строки в блоке итогов
    for current_row_num in range(summary_start_row, ws.max_row + 1):
        current_row_tuple = ws[current_row_num]
        if all(cell.value is None for cell in current_row_tuple):
            break  # Пустая строка означает конец блока

        first_cell_in_row = current_row_tuple[0]

        log.debug(
            f"get_summary: Обработка summary-строки {current_row_num} " f"со значением '{first_cell_in_row.value}'"
        )

        summary_label_raw = str(first_cell_in_row.value).strip().lower() if first_cell_in_row.value is not None else ""
        summary_key = f"merged_{current_row_num}"

        if "итого" in summary_label_raw and "ндс" in summary_label_raw:
            summary_key = JSON_KEY_TOTAL_COST_VAT
        elif "в том числе ндс" in summary_label_raw or (
            "ндс" in summary_label_raw and "итого" not in summary_label_raw
        ):
            summary_key = JSON_KEY_VAT
        elif TABLE_PARSE_DEVIATION_FROM_CALCULATED_COST in summary_label_raw:
            summary_key = JSON_KEY_DEVIATION_FROM_CALCULATED_COST
        elif TABLE_PARSE_INITIAL_COST in summary_label_raw:
            summary_key = JSON_KEY_INITIAL_COST

        summary_item_data: Dict[str, Any] = {
            JSON_KEY_JOB_TITLE: first_cell_in_row.value,
        }
        contractor_specific_summary_data = parse_contractor_row(ws, current_row_num, contractor)
        summary_item_data.update(contractor_specific_summary_data)

        summary[summary_key] = summary_item_data

    return summary
