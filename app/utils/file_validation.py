# app/utils/file_validation.py
import logging
import zipfile
from io import BytesIO
from typing import Optional

from fastapi import HTTPException, UploadFile
from fastapi.concurrency import run_in_threadpool
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# --- Logger ---
logger = logging.getLogger(__name__)

# --- лимиты (можно вынести в .env при желании) ---
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_UNZIPPED_SIZE = 200 * 1024 * 1024  # 200 MB (anti zip-bomb)
MAX_ZIP_ENTRIES = 5000  # макс. число файлов в архиве
READ_CHUNK_SIZE = 512 * 1024  # 512 KB

MAX_SHEETS = 1  # ровно один лист
MAX_ROWS_PER_SHEET = 5000  # не более 5000 строк на листе

ALLOWED_EXTS = {".xlsx"}  # при необходимости добавьте ".xlsm"


def _ext_ok(filename: Optional[str]) -> bool:
    if not filename:
        return False
    name = filename.lower()
    return any(name.endswith(ext) for ext in ALLOWED_EXTS)


async def _read_limited(upload_file: UploadFile) -> bytes:
    """Читает файл кусками и ограничивает общий размер."""
    total = 0
    chunks = []
    while True:
        chunk = await upload_file.read(READ_CHUNK_SIZE)
        if not chunk:
            break
        total += len(chunk)
        if total > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"Файл слишком большой. Максимум {MAX_FILE_SIZE // (1024*1024)} MB.",
            )
        chunks.append(chunk)
    return b"".join(chunks)


def _zip_guard(xlsx_bytes: bytes) -> None:
    """Проверка структуры XLSX как ZIP (anti zip-bomb) перед openpyxl."""
    try:
        with zipfile.ZipFile(BytesIO(xlsx_bytes)) as zf:
            infos = zf.infolist()
            if len(infos) == 0 or len(infos) > MAX_ZIP_ENTRIES:
                raise HTTPException(status_code=400, detail="Некорректный XLSX-архив (подозрительная структура).")
            total_unzipped = 0
            for i in infos:
                if i.file_size < 0 or i.compress_size < 0:
                    raise HTTPException(status_code=400, detail="Некорректный XLSX-архив (испорчённые размеры файлов).")
                total_unzipped += i.file_size
                if total_unzipped > MAX_UNZIPPED_SIZE:
                    raise HTTPException(
                        status_code=400, detail="Слишком большой распакованный размер (возможна zip-бомба)."
                    )
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Файл не является валидным XLSX (повреждённый архив).")


def _openpyxl_quick_checks(xlsx_bytes: bytes) -> None:
    """
    Синхронная часть (запускается в threadpool):
    - Проверка валидности XLSX (openpyxl).
    - Быстрая проверка формы: ровно 1 лист и не более 5000 строк.
    """
    logger = logging.getLogger(__name__)

    wb = None
    try:
        logger.info("📊 Loading workbook with openpyxl...")
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        sheetnames = wb.sheetnames
        logger.info("📋 Found sheets: %s", sheetnames)

        if not sheetnames:
            raise HTTPException(status_code=400, detail="В книге нет листов.")
        if len(sheetnames) != MAX_SHEETS:
            raise HTTPException(status_code=400, detail=f"В книге должен быть ровно {MAX_SHEETS} лист.")

        ws = wb[sheetnames[0]]

        # В read_only режиме max_row может возвращать максимальное значение Excel (1048576)
        # Вместо этого подсчитаем реальные строки с данными
        actual_rows = 0
        for row in ws.iter_rows(max_row=MAX_ROWS_PER_SHEET + 1):
            # Проверяем, есть ли хотя бы одна непустая ячейка в строке
            if any(cell.value is not None for cell in row):
                actual_rows += 1
                # Прерываем, если превысили лимит
                if actual_rows > MAX_ROWS_PER_SHEET:
                    break

        logger.info("📏 Sheet has %d rows with data (max allowed: %d)", actual_rows, MAX_ROWS_PER_SHEET)

        if actual_rows > MAX_ROWS_PER_SHEET:
            raise HTTPException(
                status_code=400,
                detail=f"Слишком много строк: {actual_rows}. Допустимо не более {MAX_ROWS_PER_SHEET}.",
            )
    except InvalidFileException as e:
        logger.exception("❌ InvalidFileException")
        raise HTTPException(status_code=400, detail="Файл не является валидным Excel-файлом.") from e
    except HTTPException:
        # прокидываем наши осмысленные ошибки как есть
        raise
    except Exception as e:
        logger.exception("❌ Unexpected error in _openpyxl_quick_checks")
        raise HTTPException(
            status_code=400,
            detail="Ошибка при проверке Excel-файла. Убедитесь, что файл не поврежден.",
        ) from e
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


async def validate_excel_upload_file(upload_file: UploadFile) -> bytes:
    """
    Валидирует загруженный Excel-файл (.xlsx) перед дальнейшей обработкой.

    Проверки:
      1) Расширение (ALLOWED_EXTS).
      2) Чтение файла кусками (MAX_FILE_SIZE) — HTTP 413 при превышении.
      3) Anti ZIP-bomb: структура ZIP, число записей, суммарный uncompressed size.
      4) Проверка валидности openpyxl в threadpool.
      5) Быстрая структурная проверка: ровно 1 лист и ≤ MAX_ROWS_PER_SHEET строк.

    Возвращает:
      bytes — содержимое файла.
    """
    logger.info("🔍 VALIDATION DEBUG: filename=%s, content_type=%s", upload_file.filename, upload_file.content_type)

    if not _ext_ok(upload_file.filename):
        allowed = ", ".join(sorted(ALLOWED_EXTS))
        logger.error("❌ Extension check failed: filename=%s, allowed=%s", upload_file.filename, allowed)
        raise HTTPException(status_code=400, detail=f"Файл должен быть в формате: {allowed}")

    try:
        file_bytes = await _read_limited(upload_file)
        logger.info("✅ File read successfully: size=%d bytes", len(file_bytes))
    except HTTPException:
        logger.error("❌ File read failed with HTTPException")
        raise
    except Exception:
        logger.exception("❌ File read failed with generic exception")
        raise HTTPException(status_code=500, detail="Не удалось прочитать файл.")

    try:
        _zip_guard(file_bytes)
        logger.info("✅ ZIP structure validation passed")
    except HTTPException:
        logger.error("❌ ZIP structure validation failed")
        raise

    try:
        # openpyxl и структурные проверки — в threadpool, чтобы не блокировать event loop
        await run_in_threadpool(_openpyxl_quick_checks, file_bytes)
        logger.info("✅ OpenPyXL validation passed")
    except HTTPException as e:
        logger.exception("❌ OpenPyXL validation failed: %s", e.detail)
        raise
    except Exception as e:
        logger.exception("❌ OpenPyXL validation failed with unexpected error")
        raise HTTPException(status_code=400, detail=f"Ошибка валидации Excel: {e!s}") from e

    logger.info("🎉 All validations passed successfully")
    return file_bytes
