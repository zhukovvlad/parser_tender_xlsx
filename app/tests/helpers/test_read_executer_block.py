# tests/test_read_executer_block.py

import pytest
from openpyxl import Workbook

# Импортируем константы из проекта
from app.constants import (
    JSON_KEY_EXECUTOR_NAME,
    JSON_KEY_EXECUTOR_PHONE, 
    JSON_KEY_EXECUTOR_DATE,
    TABLE_PARSE_EXECUTOR,
    TABLE_PARSE_TELEPHONE,
    TABLE_PARSE_PREPARATION_DATE,
)

# Импортируем тестируемую функцию
from app.helpers.read_executer_block import read_executer_block


def set_max_row(ws, row_num):
    """Хелпер для явного установления максимальной строки в тестах."""
    ws[f"A{row_num}"] = "dummy_data_to_set_max_row"


# --- ОБНОВЛЕННЫЕ И НОВЫЕ ТЕСТЫ ---


def test_read_executer_block_happy_path_with_colon():
    """
    Проверяет идеальный сценарий, когда дата содержит двоеточие-разделитель.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    ws["B15"] = "Исполнитель: Иванов И.И."
    ws["B16"] = "Телефон: +7 (999) 123-45-67"
    ws["B17"] = "Дата составления: 15.07.2025"

    # Act
    result = read_executer_block(ws)

    # Assert
    assert result[JSON_KEY_EXECUTOR_NAME] == "Иванов И.И."
    assert result[JSON_KEY_EXECUTOR_PHONE] == "+7 (999) 123-45-67"
    # ОБНОВЛЕНО: Проверяем, что дата извлечена без двоеточия
    assert result[JSON_KEY_EXECUTOR_DATE] == "15.07.2025"


def test_read_executer_block_date_format_without_colon():
    """
    НОВЫЙ ТЕСТ: Проверяет формат даты без двоеточия-разделителя,
    но с двоеточиями во времени.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    ws["B17"] = "Дата составления 07.05.2025 18:49:35"

    # Act
    result = read_executer_block(ws)

    # Assert
    assert result[JSON_KEY_EXECUTOR_NAME] is None
    assert result[JSON_KEY_EXECUTOR_DATE] == "07.05.2025 18:49:35"


def test_read_executer_block_date_format_with_space_before_colon():
    """
    НОВЫЙ ТЕСТ: Проверяет, что логика справляется с пробелом между
    ключевым словом и двоеточием.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    ws["B17"] = "Дата составления : 12.12.2025"

    # Act
    result = read_executer_block(ws)

    # Assert
    assert result[JSON_KEY_EXECUTOR_DATE] == "12.12.2025"


def test_read_executer_block_is_case_insensitive():
    """
    Проверяет, что поиск ключевых слов нечувствителен к регистру для всех полей.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 10)  # Сканируем 5, 6, 7
    ws["B5"] = "ИСПОЛНИТЕЛЬ: Петров П.П."
    ws["B6"] = "телефон: 88005553535"
    ws["B7"] = "ДАТА СОСТАВЛЕНИЯ: 01.01.2025"

    # Act
    result = read_executer_block(ws)

    # Assert
    assert result[JSON_KEY_EXECUTOR_NAME] == "Петров П.П."
    assert result[JSON_KEY_EXECUTOR_PHONE] == "88005553535"
    # ОБНОВЛЕНО: Проверяем, что дата извлечена без двоеточия
    assert result[JSON_KEY_EXECUTOR_DATE] == "01.01.2025"


# --- ОСТАЛЬНЫЕ ТЕСТЫ ОСТАЮТСЯ БЕЗ ИЗМЕНЕНИЙ, ТАК КАК ОНИ ПРОВЕРЯЮТ ОБЩУЮ ЛОГИКУ ---


def test_read_executer_block_no_data_found():
    """
    Проверяет, что на листе без ключевых слов возвращаются None.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    ws["B15"] = "Просто какой-то текст"

    # Act
    result = read_executer_block(ws)

    # Assert
    assert all(value is None for value in result.values())


def test_read_executer_block_ignores_data_in_wrong_column():
    """
    Проверяет, что данные ищутся только во второй колонке ('B').
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    ws["A15"] = "Исполнитель: Иванов И.И."

    # Act
    result = read_executer_block(ws)

    # Assert
    assert result[JSON_KEY_EXECUTOR_NAME] is None


def test_read_executer_block_handles_small_sheet_gracefully():
    """
    Проверяет, что функция не падает на листах с малым количеством строк.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "data"
    ws["A2"] = "data"

    # Act
    try:
        result = read_executer_block(ws)
    except Exception as e:
        pytest.fail(f"Функция упала с ошибкой на маленьком листе: {e}")

    # Assert
    assert all(value is None for value in result.values())


def test_read_executer_block_ignores_non_string_values():
    """
    Проверяет, что функция игнорирует ячейки с не-строковыми данными (числа, None и т.д.).
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    ws["B15"] = 123456789  # Число вместо строки

    # Act
    result = read_executer_block(ws)

    # Assert
    assert all(value is None for value in result.values())


def test_read_executer_block_handles_missing_colon_for_name():
    """
    Проверяет, что код не падает, если у исполнителя или телефона нет двоеточия.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    ws["B15"] = "Исполнитель Иванов И.И."  # Нет двоеточия

    # Act
    result = read_executer_block(ws)

    # Assert
    # Ошибка IndexError должна быть обработана, и значение останется None
    assert result[JSON_KEY_EXECUTOR_NAME] is None


def test_read_executer_block_ignores_data_in_wrong_rows():
    """
    Проверяет, что сканируется только предопределенный диапазон строк.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)  # -> сканируем 15, 16, 17

    # Данные в строках 14 и 18 должны быть проигнорированы
    ws["B14"] = "Исполнитель: Неправильный"
    ws["B18"] = "Телефон: 555-55-55"

    # Act
    result = read_executer_block(ws)

    # Assert
    assert all(value is None for value in result.values())


def test_read_executer_block_handles_missing_colon_for_phone():
    """
    НОВЫЙ ТЕСТ: Проверяет обработку `except IndexError` для телефона.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    # Ключевое слово есть, а двоеточия нет
    ws["B16"] = "Телефон 89991234567"

    # Act
    result = read_executer_block(ws)

    # Assert
    # Код должен обработать IndexError и оставить значение None
    assert result[JSON_KEY_EXECUTOR_PHONE] is None


def test_read_executer_block_handles_missing_colon_for_phone():
    """
    Проверяет обработку `except IndexError` для телефона, если нет двоеточия.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    set_max_row(ws, 20)
    # Ключевое слово есть, а двоеточия нет
    ws["B16"] = "Телефон 89991234567"

    # Act
    result = read_executer_block(ws)

    # Assert
    # Код должен корректно обработать IndexError и оставить значение None
    assert result[JSON_KEY_EXECUTOR_PHONE] is None
