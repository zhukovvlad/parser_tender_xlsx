"""
Тесты для модуля get_proposals.

Эти тесты проверяют реальное поведение функции get_proposals,
которая агрегирует предложения всех подрядчиков для конкретного лота.

Философия тестирования:
- Тестируем ОРКЕСТРАЦИЮ: как функция координирует работу других модулей
- Проверяем АГРЕГАЦИЮ: корректность сборки данных из разных источников
- Валидируем РАЗДЕЛЕНИЕ: правильность разделения данных по лотам и общих данных
- Тестируем ИНТЕГРАЦИЮ: взаимодействие с read_contractors, get_lot_positions, get_summary, get_additional_info
"""

from unittest.mock import MagicMock, Mock, patch

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    JSON_KEY_CONTRACTOR_ACCREDITATION,
    JSON_KEY_CONTRACTOR_ADDITIONAL_INFO,
    JSON_KEY_CONTRACTOR_ADDRESS,
    JSON_KEY_CONTRACTOR_COORDINATE,
    JSON_KEY_CONTRACTOR_HEIGHT,
    JSON_KEY_CONTRACTOR_INDEX,
    JSON_KEY_CONTRACTOR_INN,
    JSON_KEY_CONTRACTOR_ITEMS,
    JSON_KEY_CONTRACTOR_POSITIONS,
    JSON_KEY_CONTRACTOR_SUMMARY,
    JSON_KEY_CONTRACTOR_TITLE,
    JSON_KEY_CONTRACTOR_WIDTH,
)
from app.excel_parser.get_proposals import get_proposals


@pytest.fixture
def empty_worksheet():
    """Создает пустой Excel worksheet для тестов."""
    wb = Workbook()
    ws = wb.active
    return ws


@pytest.fixture
def sample_contractors_data():
    """Образец данных подрядчиков от read_contractors."""
    return [
        {
            "value": "HEADER_ROW",
            "row_start": 1,
            "column_start": 1,
            "coordinate": "A1",
            "merged_shape": {"rowspan": 1, "colspan": 1},
        },
        {
            "value": "ООО Строитель",
            "row_start": 2,
            "column_start": 2,
            "coordinate": "B2",
            "merged_shape": {"rowspan": 1, "colspan": 1},
        },
        {
            "value": "АО Подрядчик",
            "row_start": 2,
            "column_start": 4,
            "coordinate": "D2",
            "merged_shape": {"rowspan": 1, "colspan": 1},
        },
    ]


@pytest.fixture
def sample_positions_data():
    """Образец данных позиций от get_lot_positions."""
    return [
        {"position": "Позиция 1", "quantity": 10, "price": 1000},
        {"position": "Позиция 2", "quantity": 5, "price": 2000},
    ]


@pytest.fixture
def sample_summary_data():
    """Образец данных итогов от get_summary."""
    return {"total_amount": 15000, "vat": 2700}


@pytest.fixture
def sample_additional_info():
    """Образец дополнительной информации от get_additional_info."""
    return {"comment": "Дополнительная информация", "deadline": "30 дней"}


class TestGetProposalsBasicBehavior:
    """Тесты базового поведения функции get_proposals."""

    def test_returns_empty_dict_when_no_contractors(self, empty_worksheet):
        """Должна возвращать пустой словарь, если подрядчиков нет."""
        with patch("app.excel_parser.get_proposals.read_contractors", return_value=None):
            result = get_proposals(empty_worksheet, 10, 20)
            assert result == {}

    def test_returns_empty_dict_when_empty_contractors_list(self, empty_worksheet):
        """Должна возвращать пустой словарь, если список подрядчиков пустой."""
        with patch("app.excel_parser.get_proposals.read_contractors", return_value=[]):
            result = get_proposals(empty_worksheet, 10, 20)
            assert result == {}

    def test_skips_first_contractor_entry(self, empty_worksheet, sample_contractors_data):
        """Должна пропускать первый элемент списка (заголовок)."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)
            # Должно быть 2 подрядчика (индексы 1 и 2), первый пропущен
            assert len(result) == 2
            assert "contractor_1" in result
            assert "contractor_2" in result


class TestGetProposalsDataExtraction:
    """Тесты извлечения данных из ячеек Excel."""

    def test_extracts_basic_contractor_info(self, empty_worksheet, sample_contractors_data):
        """Должна извлекать базовую информацию о подрядчике из ячеек."""
        # Подготавливаем данные в worksheet
        empty_worksheet.cell(row=3, column=2, value="1234567890")  # ИНН
        empty_worksheet.cell(row=4, column=2, value="г. Москва, ул. Ленина, д.1")  # Адрес
        empty_worksheet.cell(row=5, column=2, value="Аккредитация есть")  # Аккредитация

        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            contractor_1 = result["contractor_1"]
            assert contractor_1[JSON_KEY_CONTRACTOR_TITLE] == "ООО Строитель"
            assert contractor_1[JSON_KEY_CONTRACTOR_INN] == "1234567890"
            assert contractor_1[JSON_KEY_CONTRACTOR_ADDRESS] == "г. Москва, ул. Ленина, д.1"
            assert contractor_1[JSON_KEY_CONTRACTOR_ACCREDITATION] == "Аккредитация есть"

    def test_extracts_coordinate_and_dimensions(self, empty_worksheet, sample_contractors_data):
        """Должна извлекать координаты и размеры ячеек."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            contractor_1 = result["contractor_1"]
            assert contractor_1[JSON_KEY_CONTRACTOR_COORDINATE] == "B2"
            assert contractor_1[JSON_KEY_CONTRACTOR_WIDTH] == 1
            assert contractor_1[JSON_KEY_CONTRACTOR_HEIGHT] == 1

    def test_handles_merged_cells(self, empty_worksheet):
        """Должна корректно обрабатывать объединенные ячейки."""
        contractors_with_merged = [
            {
                "value": "HEADER",
                "row_start": 1,
                "column_start": 1,
                "coordinate": "A1",
                "merged_shape": {"rowspan": 1, "colspan": 1},
            },
            {
                "value": "ООО Большая компания",
                "row_start": 2,
                "column_start": 2,
                "coordinate": "B2",
                "merged_shape": {"rowspan": 2, "colspan": 3},  # Объединенная ячейка
            },
        ]

        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=contractors_with_merged),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            contractor_1 = result["contractor_1"]
            assert contractor_1[JSON_KEY_CONTRACTOR_WIDTH] == 3
            assert contractor_1[JSON_KEY_CONTRACTOR_HEIGHT] == 2
            # Для объединенных ячеек ИНН/адрес/аккредитация должны быть None
            assert contractor_1[JSON_KEY_CONTRACTOR_INN] is None
            assert contractor_1[JSON_KEY_CONTRACTOR_ADDRESS] is None
            assert contractor_1[JSON_KEY_CONTRACTOR_ACCREDITATION] is None


class TestGetProposalsModuleIntegration:
    """Тесты интеграции с другими модулями."""

    def test_calls_get_lot_positions_with_correct_parameters(self, empty_worksheet, sample_contractors_data):
        """Должна вызывать get_lot_positions с правильными параметрами лота."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions") as mock_positions,
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            mock_positions.return_value = []

            get_proposals(empty_worksheet, 15, 25)

            # Должна вызываться для каждого подрядчика с границами лота
            assert mock_positions.call_count == 2

            # Проверяем параметры первого вызова
            call_args = mock_positions.call_args_list[0]
            assert call_args[0][0] == empty_worksheet  # worksheet
            assert call_args[0][1] == sample_contractors_data[1]  # contractor_details
            assert call_args[1]["lot_start_row"] == 15
            assert call_args[1]["lot_end_row"] == 25

    def test_calls_get_summary_for_each_contractor(self, empty_worksheet, sample_contractors_data):
        """Должна вызывать get_summary для каждого подрядчика."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary") as mock_summary,
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            mock_summary.return_value = {}

            get_proposals(empty_worksheet, 10, 20)

            # Должна вызываться для каждого подрядчика (2 раза)
            assert mock_summary.call_count == 2

            # Проверяем параметры вызовов
            for i, call_args in enumerate(mock_summary.call_args_list):
                assert call_args[0][0] == empty_worksheet
                assert call_args[0][1] == sample_contractors_data[i + 1]  # +1 т.к. первый пропускается

    def test_calls_get_additional_info_for_each_contractor(self, empty_worksheet, sample_contractors_data):
        """Должна вызывать get_additional_info для каждого подрядчика."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info") as mock_additional,
        ):

            mock_additional.return_value = {}

            get_proposals(empty_worksheet, 10, 20)

            # Должна вызываться для каждого подрядчика (2 раза)
            assert mock_additional.call_count == 2


class TestGetProposalsDataAggregation:
    """Тесты агрегации данных из разных источников."""

    def test_combines_positions_and_summary_in_items(
        self, empty_worksheet, sample_contractors_data, sample_positions_data, sample_summary_data
    ):
        """Должна объединять позиции и итоги в структуру items."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=sample_positions_data),
            patch("app.excel_parser.get_proposals.get_summary", return_value=sample_summary_data),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            contractor_1 = result["contractor_1"]
            items = contractor_1[JSON_KEY_CONTRACTOR_ITEMS]

            assert items[JSON_KEY_CONTRACTOR_POSITIONS] == sample_positions_data
            assert items[JSON_KEY_CONTRACTOR_SUMMARY] == sample_summary_data

    def test_includes_additional_info_in_proposal(
        self, empty_worksheet, sample_contractors_data, sample_additional_info
    ):
        """Должна включать дополнительную информацию в предложение."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value=sample_additional_info),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            contractor_1 = result["contractor_1"]
            assert contractor_1[JSON_KEY_CONTRACTOR_ADDITIONAL_INFO] == sample_additional_info

    def test_creates_complete_proposal_structure(
        self,
        empty_worksheet,
        sample_contractors_data,
        sample_positions_data,
        sample_summary_data,
        sample_additional_info,
    ):
        """Должна создавать полную структуру предложения со всеми полями."""
        # Подготавливаем данные в worksheet
        empty_worksheet.cell(row=3, column=2, value="1234567890")  # ИНН
        empty_worksheet.cell(row=4, column=2, value="г. Москва")  # Адрес
        empty_worksheet.cell(row=5, column=2, value="Есть")  # Аккредитация

        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=sample_positions_data),
            patch("app.excel_parser.get_proposals.get_summary", return_value=sample_summary_data),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value=sample_additional_info),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            contractor_1 = result["contractor_1"]

            # Проверяем все ожидаемые поля
            expected_fields = [
                JSON_KEY_CONTRACTOR_TITLE,
                JSON_KEY_CONTRACTOR_INN,
                JSON_KEY_CONTRACTOR_ADDRESS,
                JSON_KEY_CONTRACTOR_ACCREDITATION,
                JSON_KEY_CONTRACTOR_COORDINATE,
                JSON_KEY_CONTRACTOR_WIDTH,
                JSON_KEY_CONTRACTOR_HEIGHT,
                JSON_KEY_CONTRACTOR_ITEMS,
                JSON_KEY_CONTRACTOR_ADDITIONAL_INFO,
            ]

            for field in expected_fields:
                assert field in contractor_1, f"Отсутствует поле {field}"

            # Проверяем структуру items
            items = contractor_1[JSON_KEY_CONTRACTOR_ITEMS]
            assert JSON_KEY_CONTRACTOR_POSITIONS in items
            assert JSON_KEY_CONTRACTOR_SUMMARY in items


class TestGetProposalsMultipleContractors:
    """Тесты работы с несколькими подрядчиками."""

    def test_processes_multiple_contractors(self, empty_worksheet, sample_contractors_data):
        """Должна обрабатывать несколько подрядчиков корректно."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            assert len(result) == 2
            assert "contractor_1" in result
            assert "contractor_2" in result

            # Проверяем, что названия подрядчиков правильные
            assert result["contractor_1"][JSON_KEY_CONTRACTOR_TITLE] == "ООО Строитель"
            assert result["contractor_2"][JSON_KEY_CONTRACTOR_TITLE] == "АО Подрядчик"

    def test_contractor_keys_increment_correctly(self, empty_worksheet):
        """Должна корректно нумеровать ключи подрядчиков."""
        # Создаем данные с 4 подрядчиками (первый - заголовок)
        contractors_data = [
            {
                "value": "HEADER",
                "row_start": 1,
                "column_start": 1,
                "coordinate": "A1",
                "merged_shape": {"rowspan": 1, "colspan": 1},
            },
            {
                "value": "Подрядчик 1",
                "row_start": 2,
                "column_start": 2,
                "coordinate": "B2",
                "merged_shape": {"rowspan": 1, "colspan": 1},
            },
            {
                "value": "Подрядчик 2",
                "row_start": 2,
                "column_start": 4,
                "coordinate": "D2",
                "merged_shape": {"rowspan": 1, "colspan": 1},
            },
            {
                "value": "Подрядчик 3",
                "row_start": 2,
                "column_start": 6,
                "coordinate": "F2",
                "merged_shape": {"rowspan": 1, "colspan": 1},
            },
        ]

        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            assert len(result) == 3
            assert "contractor_1" in result
            assert "contractor_2" in result
            assert "contractor_3" in result


class TestGetProposalsEdgeCases:
    """Тесты граничных случаев и обработки ошибок."""

    def test_handles_missing_contractor_fields(self, empty_worksheet):
        """Должна обрабатывать отсутствующие поля в данных подрядчика."""
        incomplete_contractors = [
            {
                "value": "HEADER",
                "row_start": 1,
                "column_start": 1,
                "coordinate": "A1",
                "merged_shape": {"rowspan": 1, "colspan": 1},
            },
            {
                "value": "Неполные данные",
                # Отсутствуют некоторые поля
                "coordinate": "B2",
            },
        ]

        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=incomplete_contractors),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            # Функция должна обработать это без ошибок
            assert len(result) == 1
            contractor_1 = result["contractor_1"]
            assert contractor_1[JSON_KEY_CONTRACTOR_TITLE] == "Неполные данные"
            assert contractor_1[JSON_KEY_CONTRACTOR_COORDINATE] == "B2"
            # Недостающие размеры должны быть по умолчанию
            assert contractor_1[JSON_KEY_CONTRACTOR_WIDTH] == 1  # default colspan
            assert contractor_1[JSON_KEY_CONTRACTOR_HEIGHT] == 1  # default rowspan

    def test_handles_none_values_in_cells(self, empty_worksheet, sample_contractors_data):
        """Должна обрабатывать None значения в ячейках Excel."""
        # Ячейки остаются пустыми (None значения)

        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            result = get_proposals(empty_worksheet, 10, 20)

            contractor_1 = result["contractor_1"]
            assert contractor_1[JSON_KEY_CONTRACTOR_INN] is None
            assert contractor_1[JSON_KEY_CONTRACTOR_ADDRESS] is None
            assert contractor_1[JSON_KEY_CONTRACTOR_ACCREDITATION] is None

    def test_handles_module_function_exceptions(self, empty_worksheet, sample_contractors_data):
        """Должна обрабатывать исключения от вызываемых функций модулей."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", side_effect=Exception("Ошибка позиций")),
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            # Функция должна пропустить исключение вверх
            with pytest.raises(Exception, match="Ошибка позиций"):
                get_proposals(empty_worksheet, 10, 20)


class TestGetProposalsLotBoundaries:
    """Тесты корректности передачи границ лота."""

    def test_passes_lot_boundaries_to_positions_function(self, empty_worksheet, sample_contractors_data):
        """Должна передавать правильные границы лота в get_lot_positions."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions") as mock_positions,
            patch("app.excel_parser.get_proposals.get_summary", return_value={}),
            patch("app.excel_parser.get_proposals.get_additional_info", return_value={}),
        ):

            mock_positions.return_value = []

            # Тестируем с разными границами
            get_proposals(empty_worksheet, 100, 200)

            # Проверяем, что все вызовы получили правильные границы
            for call_args in mock_positions.call_args_list:
                assert call_args[1]["lot_start_row"] == 100
                assert call_args[1]["lot_end_row"] == 200

    def test_lot_boundaries_do_not_affect_summary_and_additional_info(self, empty_worksheet, sample_contractors_data):
        """Границы лота не должны влиять на get_summary и get_additional_info."""
        with (
            patch("app.excel_parser.get_proposals.read_contractors", return_value=sample_contractors_data),
            patch("app.excel_parser.get_proposals.get_lot_positions", return_value=[]),
            patch("app.excel_parser.get_proposals.get_summary") as mock_summary,
            patch("app.excel_parser.get_proposals.get_additional_info") as mock_additional,
        ):

            mock_summary.return_value = {}
            mock_additional.return_value = {}

            get_proposals(empty_worksheet, 100, 200)

            # Эти функции должны вызываться только с worksheet и contractor_details
            for call_args in mock_summary.call_args_list:
                assert len(call_args[0]) == 2  # только ws и contractor_details
                assert len(call_args[1]) == 0  # нет дополнительных параметров

            for call_args in mock_additional.call_args_list:
                assert len(call_args[0]) == 2  # только ws и contractor_details
                assert len(call_args[1]) == 0  # нет дополнительных параметров
