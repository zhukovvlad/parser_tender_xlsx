# tests/helpers/test_read_headers.py

import pytest
from openpyxl import Workbook

# 1. Импортируем РЕАЛЬНЫЕ константы, которые использует ваша функция
from app.constants import (
    JSON_KEY_TENDER_ADDRESS,
    JSON_KEY_TENDER_ID,
    JSON_KEY_TENDER_OBJECT,
    JSON_KEY_TENDER_TITLE,
    TABLE_PARSE_ADDRESS,
    TABLE_PARSE_OBJECT,
    TABLE_PARSE_TENDER_SUBJECT,
)
from app.excel_parser.read_headers import read_headers


# Фикстура для мока зависимости sanitize_text
@pytest.fixture
def mock_sanitize_text(mocker):
    """Мокает функцию sanitize_text, чтобы она просто возвращала входные данные."""
    mocker.patch("app.excel_parser.read_headers.sanitize_text", side_effect=lambda text: text)


# --- Существующие и улучшенные тесты ---


def test_read_headers_happy_path(mock_sanitize_text):
    """
    Проверяет идеальный сценарий: все данные на месте и корректны.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A3"] = TABLE_PARSE_TENDER_SUBJECT
    ws["B3"] = "№ID-123 Закупка оборудования"
    ws["A4"] = TABLE_PARSE_OBJECT
    ws["C4"] = "Главный корпус"  # Значение может быть в любой колонке
    ws["A5"] = TABLE_PARSE_ADDRESS
    ws["B5"] = "г. Тест, ул. Тестовая, д. 1"

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert extracted_data[JSON_KEY_TENDER_ID] == "ID-123"
    assert extracted_data[JSON_KEY_TENDER_TITLE] == "Закупка оборудования"
    assert extracted_data[JSON_KEY_TENDER_OBJECT] == "Главный корпус"
    assert extracted_data[JSON_KEY_TENDER_ADDRESS] == "г. Тест, ул. Тестовая, д. 1"


def test_read_headers_handles_empty_sheet(mock_sanitize_text):
    """
    Проверяет, что на пустом листе функция не падает и возвращает None для всех полей.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert all(value is None for value in extracted_data.values())


def test_read_headers_subject_with_only_id(mock_sanitize_text):
    """
    Проверяет логику, когда в предмете тендера есть только ID, без названия.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A3"] = TABLE_PARSE_TENDER_SUBJECT
    ws["B3"] = "№456789"  # Нет пробела и названия после номера

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert extracted_data[JSON_KEY_TENDER_ID] == "456789"
    assert extracted_data[JSON_KEY_TENDER_TITLE] == "456789"  # ID используется как название


def test_read_headers_ignores_data_outside_scan_range(mock_sanitize_text):
    """
    Проверяет, что данные в строках, не входящих в диапазон 3-5, игнорируются.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A2"] = TABLE_PARSE_TENDER_SUBJECT
    ws["B2"] = "№ID-WRONG Закупка"
    ws["A6"] = TABLE_PARSE_OBJECT
    ws["B6"] = "Неправильный корпус"

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert all(value is None for value in extracted_data.values())


def test_read_headers_handles_messy_data_and_extra_spaces(mock_sanitize_text):
    """
    Проверяет, что функция справляется с лишними пробелами и пустыми ячейками.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A4"] = "  Объект  "  # Лишние пробелы в ключе
    ws["C4"] = "   Здание АБК   "  # Лишние пробелы в значении

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert extracted_data[JSON_KEY_TENDER_OBJECT] == "Здание АБК"


def test_read_headers_ignores_keyword_if_not_first_element(mock_sanitize_text):
    """
    Проверяет, что ключевое слово игнорируется, если оно не является
    первым непустым текстом в строке.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A4"] = "Дополнительно:"
    ws["B4"] = TABLE_PARSE_OBJECT
    ws["C4"] = "Какой-то корпус"

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert extracted_data[JSON_KEY_TENDER_OBJECT] is None


def test_read_headers_parses_subject_split_by_first_space(mock_sanitize_text):
    """
    Проверяет, что разбор "Предмета тендера" происходит по первому пробелу,
    демонстрируя текущую логику парсинга.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A3"] = TABLE_PARSE_TENDER_SUBJECT
    # Этот тест показывает, как текущая логика split(" ", 1) разберет сложную строку
    ws["B3"] = '№777-ABC"Закупка ПО "Альфа""'

    # Act
    extracted_data = read_headers(ws)

    # Assert
    # Логика делит по первому пробелу, что приводит к такому результату:
    assert extracted_data[JSON_KEY_TENDER_ID] == '777-ABC"Закупка'
    assert extracted_data[JSON_KEY_TENDER_TITLE] == 'ПО "Альфа""'


# --- Новые тесты на основе рекомендаций ---


def test_read_headers_with_partial_data(mock_sanitize_text):
    """
    Проверяет, что функция корректно работает, если найдены не все ключи.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A4"] = TABLE_PARSE_OBJECT
    ws["B4"] = "Только объект"

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert extracted_data[JSON_KEY_TENDER_ID] is None
    assert extracted_data[JSON_KEY_TENDER_TITLE] is None
    assert extracted_data[JSON_KEY_TENDER_OBJECT] == "Только объект"
    assert extracted_data[JSON_KEY_TENDER_ADDRESS] is None


def test_read_headers_key_with_empty_or_whitespace_value(mock_sanitize_text):
    """
    Проверяет, что если ключ найден, а значение пустое или из пробелов,
    результат для этого поля будет None.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A4"] = TABLE_PARSE_OBJECT
    ws["B4"] = "    "  # Значение - только пробелы
    ws["A5"] = TABLE_PARSE_ADDRESS  # Ключ есть, но дальше в строке пусто

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert extracted_data[JSON_KEY_TENDER_OBJECT] is None
    assert extracted_data[JSON_KEY_TENDER_ADDRESS] is None


def test_read_headers_duplicate_key_uses_last_one(mock_sanitize_text):
    """
    Проверяет, что при дублировании ключа используется значение из последней
    найденной строки.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A4"] = TABLE_PARSE_OBJECT
    ws["B4"] = "Старый объект"
    ws["A5"] = TABLE_PARSE_OBJECT  # Тот же ключ в строке ниже
    ws["B5"] = "Новый объект"

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert extracted_data[JSON_KEY_TENDER_OBJECT] == "Новый объект"


def test_read_headers_handles_numeric_values(mock_sanitize_text):
    """
    Проверяет, что числовые значения в ячейках корректно преобразуются в строки.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A3"] = TABLE_PARSE_TENDER_SUBJECT
    ws["B3"] = 123456789  # Число, а не строка

    # Act
    extracted_data = read_headers(ws)

    # Assert
    assert extracted_data[JSON_KEY_TENDER_ID] == "123456789"
    assert extracted_data[JSON_KEY_TENDER_TITLE] == "123456789"


def test_read_headers_subject_with_only_symbol(mock_sanitize_text):
    """
    Проверяет поведение, если в значении предмета тендера только символ '№'.
    """
    # Arrange
    wb = Workbook()
    ws = wb.active
    ws["A3"] = TABLE_PARSE_TENDER_SUBJECT
    ws["B3"] = "№"

    # Act
    extracted_data = read_headers(ws)

    # Assert
    # .replace("№", "") даст пустую строку, .strip() оставит ее пустой
    # id_candidate будет '', что приравнивается к False, поэтому None
    assert extracted_data[JSON_KEY_TENDER_ID] is None
    assert extracted_data[JSON_KEY_TENDER_TITLE] is None
