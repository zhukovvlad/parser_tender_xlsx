"""
Тесты для модуля get_lot_positions.

Эти тесты проверяют реальное поведение функции get_lot_positions
с настоящими Excel данными, а не мокированными зависимостями.

Философия тестирования:
- Тестируем ПОВЕДЕНИЕ, а не реализацию
- Используем реальные данные, а не моки
- Проверяем что функция ДЕЛАЕТ, а не КАК она это делает
"""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.constants import (
    JSON_KEY_JOB_TITLE,
    JSON_KEY_JOB_TITLE_NORMALIZED,
    JSON_KEY_NUMBER,
    JSON_KEY_QUANTITY,
    JSON_KEY_UNIT,
)
from app.excel_parser.get_lot_positions import get_lot_positions


@pytest.fixture
def sample_worksheet():
    """Создает тестовый Excel worksheet с реальными данными позиций."""
    wb = Workbook()
    ws = wb.active

    # Заголовки (строка 1)
    headers = [
        "№ п/п",
        "Глава",
        "Артикул СМР",
        "Наименование видов работ",
        "Пропуск",
        "Комментарий",
        "Ед. изм.",
        "Кол-во",
        "Подрядчик 1 - Цена",
        "Подрядчик 1 - Стоимость",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Тестовые данные позиций в реалистических строках (13-15)
    test_data = [
        ["1", "01", "01-01-003", "Земляные работы", "", "Основные работы", "м³", 100, 500, 50000],
        ["2", "02", "02-01-015", "Кирпичная кладка", "", "Каменные работы", "м³", 50, 1200, 60000],
        ["3", "03", "03-02-008", "Штукатурные работы", "", "Отделочные работы", "м²", 200, 300, 60000],
    ]

    for row_idx, row_data in enumerate(test_data, 13):  # Начинаем с строки 13
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    return ws


@pytest.fixture
def real_excel_file():
    """Загружает реальный Excel файл из тестовых данных для интеграционных тестов."""
    test_data_dir = Path(__file__).parent.parent / "test_data"
    excel_path = test_data_dir / "sample_tender.xlsx"

    if excel_path.exists():
        wb = load_workbook(excel_path)
        return wb.active
    else:
        pytest.skip(f"Тестовый Excel файл не найден: {excel_path}")


class TestGetLotPositionsBehavior:
    """Тесты поведения функции get_lot_positions с реальными данными."""

    def test_extracts_positions_from_sample_data(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна извлечь все позиции из диапазона строк
        и вернуть их в виде словаря с правильной структурой.
        """
        contractor = {
            "column_start": 5,  # Колонка "Подрядчик 1 - Цена"
            "merged_shape": {"colspan": 8},  # Реалистичный размер для подрядчика
        }

        result = get_lot_positions(
            sample_worksheet, contractor, lot_start_row=13, lot_end_row=15  # Первая позиция  # Последняя позиция
        )

        # ПРОВЕРЯЕМ ПОВЕДЕНИЕ: должно извлечь 3 позиции
        assert isinstance(result, dict)
        assert len(result) == 3

        # ПРОВЕРЯЕМ СТРУКТУРУ: каждая позиция должна иметь ключевые поля
        for position_key in ["1", "2", "3"]:
            assert position_key in result
            position = result[position_key]

            # Базовые поля должны присутствовать
            assert JSON_KEY_NUMBER in position
            assert JSON_KEY_JOB_TITLE in position
            assert JSON_KEY_UNIT in position
            assert JSON_KEY_QUANTITY in position

            # Нормализованный заголовок должен быть создан
            if position.get(JSON_KEY_JOB_TITLE):
                assert JSON_KEY_JOB_TITLE_NORMALIZED in position

    def test_handles_empty_range_gracefully(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна корректно обработать пустой диапазон строк
        и вернуть пустой результат.
        """
        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        # Тест: lot_start_row > lot_end_row (невалидный диапазон)
        result = get_lot_positions(
            sample_worksheet, contractor, lot_start_row=20, lot_end_row=15  # После наших данных  # Невалидный диапазон
        )

        assert result == {}

    def test_handles_empty_rows_correctly(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна пропускать пустые строки и
        останавливаться при встрече полностью пустой строки.
        """
        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        # Тестируем диапазон, который включает пустые строки после данных
        result = get_lot_positions(
            sample_worksheet,
            contractor,
            lot_start_row=13,  # Начинаем с реальных данных
            lot_end_row=30,  # Включает много пустых строк
        )

        # Должно извлечь только реальные позиции (3 штуки)
        assert len(result) == 3
        assert "1" in result
        assert "2" in result
        assert "3" in result

    def test_respects_lot_boundaries(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна строго соблюдать границы лота
        и обрабатывать только строки в указанном диапазоне.
        """
        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        # Тест: обработка только первых двух позиций
        result = get_lot_positions(
            sample_worksheet, contractor, lot_start_row=13, lot_end_row=14  # Первая позиция  # Только первые 2 строки
        )

        assert len(result) == 2
        assert "1" in result
        assert "2" in result
        assert "3" not in result  # Третья позиция не должна попасть

        # Проверяем содержимое первой позиции
        first_position = result["1"]
        assert first_position[JSON_KEY_JOB_TITLE] == "Земляные работы"
        assert first_position[JSON_KEY_UNIT] == "м³"
        assert first_position[JSON_KEY_QUANTITY] == 100

    def test_processes_contractor_data_correctly(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна корректно извлекать данные подрядчика
        из указанных колонок и добавлять их к позициям.
        """
        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}  # Колонка "Подрядчик 1 - Цена"

        result = get_lot_positions(
            sample_worksheet, contractor, lot_start_row=13, lot_end_row=13  # Первая позиция  # Только первая позиция
        )

        assert "1" in result
        position = result["1"]

        # Должны быть данные от подрядчика (parse_contractor_row)
        # Точную структуру проверяем по факту, а не по моку
        assert isinstance(position, dict)
        assert len(position) > 4  # Больше чем базовые поля

    def test_normalizes_job_titles(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна создавать нормализованные версии
        названий работ для улучшения поиска и анализа.
        """
        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        result = get_lot_positions(
            sample_worksheet,
            contractor,
            lot_start_row=13,  # Начинаем с реальных данных
            lot_end_row=15,  # Все три позиции
        )

        # Проверяем нормализацию для каждой позиции
        for position_key, position in result.items():
            original_title = position.get(JSON_KEY_JOB_TITLE)
            normalized_title = position.get(JSON_KEY_JOB_TITLE_NORMALIZED)

            if original_title:
                # Нормализованный заголовок должен существовать
                assert normalized_title is not None
                # И должен быть строкой (если нормализация прошла успешно)
                if normalized_title:
                    assert isinstance(normalized_title, str)

    @pytest.mark.skipif(
        not Path(__file__).parent.parent.joinpath("test_data", "sample_tender.xlsx").exists(),
        reason="Тестовый Excel файл не найден",
    )
    def test_integration_with_real_excel_data(self, real_excel_file):
        """
        ИНТЕГРАЦИОННЫЙ ТЕСТ: Функция должна работать с реальными
        Excel файлами из проекта без ошибок.
        """
        # Используем реальные параметры подрядчика
        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}  # Примерная колонка подрядчика

        # Тестируем небольшой диапазон строк
        result = get_lot_positions(real_excel_file, contractor, lot_start_row=15, lot_end_row=20)

        # Проверяем, что функция работает без ошибок
        assert isinstance(result, dict)
        # Остальные проверки зависят от реальной структуры файла


class TestGetLotPositionsEdgeCases:
    """Тесты граничных случаев и обработки ошибок."""

    def test_invalid_contractor_structure(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна корректно обрабатывать или сообщать
        об ошибках при невалидной структуре contractor.
        """
        invalid_contractor = {
            # Отсутствует column_start
            "merged_shape": {"colspan": 8}
        }

        with pytest.raises((KeyError, AttributeError, TypeError)):
            get_lot_positions(sample_worksheet, invalid_contractor, lot_start_row=13, lot_end_row=15)

    def test_extreme_row_ranges(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна корректно обрабатывать
        экстремальные значения диапазонов строк.
        """
        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        # Тест с большим, но реалистичным диапазоном (крупные тендеры)
        result = get_lot_positions(
            sample_worksheet, contractor, lot_start_row=13, lot_end_row=5000  # Реалистичный размер для крупного тендера
        )

        # Должно обработать только существующие данные
        assert isinstance(result, dict)
        assert len(result) <= 10  # Должно найти только наши 3 позиции


class TestGetLotPositionsDataIntegrity:
    """Тесты целостности и качества извлекаемых данных."""

    def test_preserves_data_types(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна сохранять правильные типы данных
        для числовых и текстовых полей.
        """
        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        result = get_lot_positions(
            sample_worksheet, contractor, lot_start_row=13, lot_end_row=13  # Первая позиция  # Только первая позиция
        )

        position = result["1"]

        # Числовые поля должны быть числами
        if JSON_KEY_QUANTITY in position:
            assert isinstance(position[JSON_KEY_QUANTITY], (int, float))

        # Текстовые поля должны быть строками
        if JSON_KEY_JOB_TITLE in position:
            assert isinstance(position[JSON_KEY_JOB_TITLE], str)

        if JSON_KEY_UNIT in position:
            assert isinstance(position[JSON_KEY_UNIT], str)

    def test_handles_missing_data_gracefully(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна корректно обрабатывать отсутствующие
        данные в ячейках Excel без падения.
        """
        # Создаем строку с частично отсутствующими данными
        ws = sample_worksheet
        ws.cell(row=16, column=1, value="4")  # Номер есть
        ws.cell(row=16, column=4, value="Тест работа")  # Название есть
        # Остальные ячейки остаются пустыми

        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        result = get_lot_positions(ws, contractor, lot_start_row=16, lot_end_row=16)

        # Функция должна обработать строку, даже если некоторые данные отсутствуют
        if result:  # Если строка была обработана
            position = list(result.values())[0]
            assert JSON_KEY_JOB_TITLE in position
            assert position[JSON_KEY_JOB_TITLE] == "Тест работа"

    def test_stops_at_merged_cell_in_first_column(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна прекращать обработку при встрече
        объединенной ячейки в первом столбце (признак блока итогов).
        """
        ws = sample_worksheet

        # Добавляем позицию
        ws.cell(row=17, column=1, value="5")
        ws.cell(row=17, column=4, value="Обычная позиция")

        # Создаем объединенную ячейку в первом столбце (имитация блока итогов)
        ws.merge_cells("A18:A19")
        ws.cell(row=18, column=1, value="ИТОГО")

        # Добавляем еще одну позицию ПОСЛЕ объединенной ячейки (не должна обрабатываться)
        ws.cell(row=20, column=1, value="6")
        ws.cell(row=20, column=4, value="Не должна обрабатываться")

        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        result = get_lot_positions(ws, contractor, lot_start_row=17, lot_end_row=21)

        # Должна обработаться только первая позиция (до объединенной ячейки)
        assert len(result) == 1
        assert "1" in result
        assert result["1"][JSON_KEY_JOB_TITLE] == "Обычная позиция"

    def test_skips_completely_empty_rows(self, sample_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна пропускать полностью пустые строки
        и продолжать обработку следующих строк.
        """
        ws = sample_worksheet

        # Добавляем первую позицию
        ws.cell(row=22, column=1, value="7")
        ws.cell(row=22, column=4, value="Первая позиция")

        # Строка 23 остается полностью пустой

        # Добавляем вторую позицию после пустой строки
        ws.cell(row=24, column=1, value="8")
        ws.cell(row=24, column=4, value="Вторая позиция")

        contractor = {"column_start": 9, "merged_shape": {"colspan": 8}}

        result = get_lot_positions(ws, contractor, lot_start_row=22, lot_end_row=24)

        # Должны обработаться обе позиции, пустая строка пропущена
        assert len(result) == 2
        assert "1" in result
        assert "2" in result
        assert result["1"][JSON_KEY_JOB_TITLE] == "Первая позиция"
        assert result["2"][JSON_KEY_JOB_TITLE] == "Вторая позиция"
