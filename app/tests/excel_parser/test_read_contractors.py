"""
Тесты для модуля read_contractors.

Эти тесты проверяют реальное поведение функции read_contractors,
которая ищет и извлекает информацию о заголовках контрагентов из Excel.

Философия тестирования:
- Тестируем ПОВЕДЕНИЕ: как функция находит строки с контрагентами
- Проверяем ПОИСК: корректность обнаружения маркерной строки
- Валидируем ДАННЫЕ: правильность извлекаемой информации о ячейках
"""

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.constants import TABLE_PARSE_CONTRACTOR_TITLE
from app.excel_parser.read_contractors import read_contractors


@pytest.fixture
def empty_worksheet():
    """Создает пустой Excel worksheet для тестов."""
    wb = Workbook()
    ws = wb.active
    return ws


@pytest.fixture
def worksheet_with_contractors():
    """Создает Excel worksheet с типичными заголовками контрагентов."""
    wb = Workbook()
    ws = wb.active

    # Добавляем некоторые обычные данные в верхние строки
    ws.cell(row=1, column=1, value="Тендерная документация")
    ws.cell(row=2, column=1, value="Объект строительства")
    ws.cell(row=3, column=1, value="Раздел работ")

    # Строка 5 - строка с заголовками контрагентов
    ws.cell(row=5, column=1, value="№ п/п")
    ws.cell(row=5, column=2, value="Наименование работ")
    ws.cell(row=5, column=3, value="Ед. изм.")
    ws.cell(row=5, column=4, value="Кол-во")
    ws.cell(row=5, column=5, value="Наименование контрагента №1")  # Маркерная ячейка
    ws.cell(row=5, column=6, value="Цена за единицу")
    ws.cell(row=5, column=7, value="Общая стоимость")
    ws.cell(row=5, column=8, value="Наименование контрагента №2")  # Второй контрагент
    ws.cell(row=5, column=9, value="Цена за единицу")
    ws.cell(row=5, column=10, value="Общая стоимость")

    return ws


@pytest.fixture
def worksheet_with_merged_cells():
    """Создает Excel worksheet с объединенными ячейками в строке контрагентов."""
    wb = Workbook()
    ws = wb.active

    # Строка 6 - строка с заголовками контрагентов и объединенными ячейками
    ws.cell(row=6, column=1, value="№ п/п")
    ws.cell(row=6, column=2, value="Наименование работ")
    ws.cell(row=6, column=3, value="Наименование контрагента ООО 'Строй'")  # Маркерная ячейка

    # Создаем объединенную ячейку для контрагента (колонки 4-6)
    ws.merge_cells("D6:F6")
    ws.cell(row=6, column=4, value="Подрядчик Альфа")

    # Еще одна обычная ячейка
    ws.cell(row=6, column=7, value="Комментарии")

    return ws


class TestReadContractorsBehavior:
    """Тесты основного поведения функции read_contractors."""

    def test_returns_none_for_empty_worksheet(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна возвращать None,
        если на листе нет строки с маркером контрагентов.
        """
        result = read_contractors(empty_worksheet)
        assert result is None

    def test_returns_none_when_no_contractor_title_found(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна возвращать None,
        если маркерный текст не найден в диапазоне строк 4-10.
        """
        ws = empty_worksheet

        # Добавляем различные данные, но без маркерного текста
        ws.cell(row=4, column=1, value="Обычные данные")
        ws.cell(row=5, column=1, value="Еще данные")
        ws.cell(row=6, column=1, value="Название работ")
        ws.cell(row=7, column=1, value="Прочая информация")

        result = read_contractors(ws)
        assert result is None

    def test_returns_list_when_contractor_title_found(self, worksheet_with_contractors):
        """
        ПОВЕДЕНИЕ: Функция должна возвращать список словарей,
        когда находит строку с маркером контрагентов.
        """
        result = read_contractors(worksheet_with_contractors)

        assert result is not None
        assert isinstance(result, list)
        assert len(result) > 0

    def test_finds_contractor_title_case_insensitive(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Поиск маркерного текста должен быть регистронезависимым.
        """
        ws = empty_worksheet

        # Тестируем разные варианты регистра
        test_cases = [
            "НАИМЕНОВАНИЕ КОНТРАГЕНТА",
            "Наименование Контрагента",
            "наименование контрагента",
            "НаИмЕнОвАнИе КоНтРаГеНтА",
        ]

        for case_variant in test_cases:
            # Очищаем лист
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

            # Добавляем маркерный текст в разном регистре
            ws.cell(row=5, column=3, value=case_variant)
            ws.cell(row=5, column=4, value="Подрядчик 1")

            result = read_contractors(ws)
            assert result is not None, f"Должен найти маркер: '{case_variant}'"
            assert len(result) == 2  # Две непустые ячейки

    def test_ignores_leading_trailing_whitespace_in_search(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна игнорировать пробелы в начале и конце
        при поиске маркерного текста.
        """
        ws = empty_worksheet

        # Добавляем маркерный текст с пробелами
        ws.cell(row=6, column=2, value="  наименование контрагента  ")
        ws.cell(row=6, column=3, value="ООО Строитель")

        result = read_contractors(ws)
        assert result is not None
        assert len(result) == 2


class TestReadContractorsSearchRange:
    """Тесты диапазона поиска (строки 4-10)."""

    def test_searches_only_in_rows_4_to_10(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна искать маркерный текст
        только в строках с 4-й по 10-ю включительно.
        """
        ws = empty_worksheet

        # Тестируем строки вне диапазона поиска
        out_of_range_cases = [
            (1, "В строке 1"),
            (2, "В строке 2"),
            (3, "В строке 3"),
            (11, "В строке 11"),
            (15, "В строке 15"),
        ]

        for row_num, description in out_of_range_cases:
            # Очищаем лист
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

            # Добавляем маркерный текст в строку вне диапазона
            ws.cell(row=row_num, column=1, value="наименование контрагента")
            ws.cell(row=row_num, column=2, value="Подрядчик")

            result = read_contractors(ws)
            assert result is None, f"Не должен найти маркер {description}"

    def test_finds_in_each_valid_row(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна находить маркерный текст
        в любой строке диапазона 4-10.
        """
        ws = empty_worksheet

        valid_rows = [4, 5, 6, 7, 8, 9, 10]

        for row_num in valid_rows:
            # Очищаем лист
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

            # Добавляем маркерный текст в текущую строку
            ws.cell(row=row_num, column=1, value="наименование контрагента")
            ws.cell(row=row_num, column=2, value="Подрядчик")

            result = read_contractors(ws)
            assert result is not None, f"Должен найти маркер в строке {row_num}"
            assert len(result) == 2

    def test_returns_first_matching_row_only(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Если маркерный текст присутствует в нескольких строках,
        функция должна вернуть данные только из первой найденной строки.
        """
        ws = empty_worksheet

        # Добавляем маркерный текст в несколько строк
        ws.cell(row=5, column=1, value="наименование контрагента")
        ws.cell(row=5, column=2, value="Первый подрядчик")

        ws.cell(row=7, column=1, value="наименование контрагента")
        ws.cell(row=7, column=2, value="Второй подрядчик")
        ws.cell(row=7, column=3, value="Третий подрядчик")

        result = read_contractors(ws)
        assert result is not None
        assert len(result) == 2  # Только из первой найденной строки (5)

        # Проверяем, что данные из строки 5
        values = [cell_info["value"] for cell_info in result]
        assert "Первый подрядчик" in values
        assert "Второй подрядчик" not in values


class TestReadContractorsDataExtraction:
    """Тесты извлечения данных из найденной строки контрагентов."""

    def test_extracts_all_non_empty_cells_from_contractor_row(self, worksheet_with_contractors):
        """
        ПОВЕДЕНИЕ: Функция должна извлечь информацию обо всех
        непустых ячейках из найденной строки контрагентов.
        """
        result = read_contractors(worksheet_with_contractors)

        assert result is not None
        # В fixture у нас 10 непустых ячеек в строке 5
        assert len(result) == 10

        # Проверяем, что все ячейки имеют значения
        for cell_info in result:
            assert cell_info["value"] is not None

    def test_skips_empty_cells_in_contractor_row(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна пропускать пустые ячейки
        в строке контрагентов.
        """
        ws = empty_worksheet

        # Создаем строку с пустыми ячейками между заполненными
        ws.cell(row=5, column=1, value="наименование контрагента")
        # Колонка 2 остается пустой
        ws.cell(row=5, column=3, value="Подрядчик 1")
        # Колонки 4-5 остаются пустыми
        ws.cell(row=5, column=6, value="Подрядчик 2")

        result = read_contractors(ws)
        assert result is not None
        assert len(result) == 3  # Только непустые ячейки

        values = [cell_info["value"] for cell_info in result]
        expected_values = ["наименование контрагента", "Подрядчик 1", "Подрядчик 2"]
        assert set(values) == set(expected_values)

    def test_cell_info_structure(self, worksheet_with_contractors):
        """
        ПОВЕДЕНИЕ: Каждый элемент в возвращаемом списке должен
        содержать правильную структуру информации о ячейке.
        """
        result = read_contractors(worksheet_with_contractors)

        required_fields = ["value", "coordinate", "column_start", "row_start"]

        for cell_info in result:
            # Проверяем обязательные поля
            for field in required_fields:
                assert field in cell_info, f"Поле {field} должно присутствовать"

            # Проверяем типы данных
            assert isinstance(cell_info["coordinate"], str)
            assert isinstance(cell_info["column_start"], int)
            assert isinstance(cell_info["row_start"], int)
            assert cell_info["column_start"] >= 1
            assert cell_info["row_start"] >= 4  # В диапазоне поиска

    def test_row_start_consistency(self, worksheet_with_contractors):
        """
        ПОВЕДЕНИЕ: Все ячейки из одной строки должны иметь
        одинаковое значение row_start.
        """
        result = read_contractors(worksheet_with_contractors)

        # Все ячейки должны быть из одной строки
        row_starts = {cell_info["row_start"] for cell_info in result}
        assert len(row_starts) == 1, "Все ячейки должны быть из одной строки"

        # И эта строка должна быть 5 (согласно fixture)
        assert list(row_starts)[0] == 5

    def test_column_start_matches_coordinate(self, worksheet_with_contractors):
        """
        ПОВЕДЕНИЕ: Значение column_start должно соответствовать
        колонке в coordinate.
        """
        result = read_contractors(worksheet_with_contractors)

        for cell_info in result:
            coordinate = cell_info["coordinate"]
            column_start = cell_info["column_start"]

            # Извлекаем букву колонки из координаты (например, "A5" -> "A")
            column_letter = "".join(filter(str.isalpha, coordinate))

            # Преобразуем букву в номер и сравниваем
            from openpyxl.utils import column_index_from_string

            expected_column = column_index_from_string(column_letter)

            assert (
                column_start == expected_column
            ), f"column_start ({column_start}) должен соответствовать coordinate ({coordinate})"


class TestReadContractorsMergedCells:
    """Тесты обработки объединенных ячеек."""

    def test_adds_merged_shape_info_for_merged_cells(self, worksheet_with_merged_cells):
        """
        ПОВЕДЕНИЕ: Для объединенных ячеек должна добавляться
        информация о merged_shape.
        """
        result = read_contractors(worksheet_with_merged_cells)

        assert result is not None

        # Ищем информацию об объединенной ячейке (D6:F6)
        merged_cell_info = None
        for cell_info in result:
            if cell_info["coordinate"] == "D6":  # Первая ячейка объединенного диапазона
                merged_cell_info = cell_info
                break

        assert merged_cell_info is not None, "Должна быть найдена объединенная ячейка D6"
        assert "merged_shape" in merged_cell_info, "Должна присутствовать информация о merged_shape"

        merged_shape = merged_cell_info["merged_shape"]
        assert isinstance(merged_shape, dict)
        assert "rowspan" in merged_shape
        assert "colspan" in merged_shape

        # D6:F6 - это 1 строка и 3 колонки
        assert merged_shape["rowspan"] == 1
        assert merged_shape["colspan"] == 3

    def test_no_merged_shape_for_regular_cells(self, worksheet_with_contractors):
        """
        ПОВЕДЕНИЕ: Для обычных (не объединенных) ячеек
        не должно быть поля merged_shape.
        """
        result = read_contractors(worksheet_with_contractors)

        for cell_info in result:
            # Обычные ячейки не должны иметь merged_shape
            assert "merged_shape" not in cell_info or cell_info.get("merged_shape") is None


class TestReadContractorsEdgeCases:
    """Тесты граничных случаев и особых ситуаций."""

    def test_handles_partial_match_in_cell_value(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна находить маркерный текст,
        когда он находится в начале строки, даже если после него есть дополнительный текст.
        """
        ws = empty_worksheet

        # Маркерный текст в начале более длинной строки
        ws.cell(row=5, column=1, value="наименование контрагента и участников тендера")
        ws.cell(row=5, column=2, value="Подрядчик")

        result = read_contractors(ws)
        assert result is not None
        assert len(result) == 2

    def test_handles_non_string_cell_values(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна корректно обрабатывать ячейки
        с не-строковыми значениями в поиске маркера.
        """
        ws = empty_worksheet

        # Добавляем различные типы данных
        ws.cell(row=5, column=1, value=123)  # Число
        ws.cell(row=5, column=2, value=None)  # Пустая ячейка
        ws.cell(row=5, column=3, value="наименование контрагента")  # Маркер
        ws.cell(row=5, column=4, value=45.67)  # Дробное число
        ws.cell(row=5, column=5, value="Подрядчик")  # Строка

        result = read_contractors(ws)
        assert result is not None
        # Должно быть 4 непустые ячейки (пропускаем None в колонке 2)
        assert len(result) == 4

        # Проверяем типы значений
        values = [cell_info["value"] for cell_info in result]
        assert 123 in values
        assert 45.67 in values
        assert "наименование контрагента" in values
        assert "Подрядчик" in values

    def test_handles_worksheet_with_no_data_in_search_range(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Функция должна корректно обрабатывать лист,
        где в диапазоне поиска (строки 4-10) нет данных.
        """
        ws = empty_worksheet

        # Добавляем данные вне диапазона поиска
        ws.cell(row=1, column=1, value="Заголовок")
        ws.cell(row=2, column=1, value="наименование контрагента")  # Вне диапазона
        ws.cell(row=15, column=1, value="Конец документа")

        result = read_contractors(ws)
        assert result is None

    def test_marker_text_must_be_at_start_of_cell_value(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Маркерный текст должен находиться в начале
        значения ячейки (после удаления пробелов).
        """
        ws = empty_worksheet

        # Маркерный текст НЕ в начале строки - не должен найтись
        ws.cell(row=5, column=1, value="Это не наименование контрагента в начале")
        ws.cell(row=5, column=2, value="Подрядчик")

        result = read_contractors(ws)
        assert result is None

        # Маркерный текст в начале строки - должен найтись
        ws.cell(row=5, column=1, value="наименование контрагента - основная информация")

        result = read_contractors(ws)
        assert result is not None


class TestReadContractorsDataIntegrity:
    """Тесты целостности и качества возвращаемых данных."""

    def test_maintains_cell_order_in_result(self, empty_worksheet):
        """
        ПОВЕДЕНИЕ: Порядок ячеек в результате должен соответствовать
        порядку колонок в Excel (слева направо).
        """
        ws = empty_worksheet

        # Заполняем ячейки в определенном порядке
        cell_data = [
            (1, "наименование контрагента"),
            (3, "Подрядчик A"),
            (5, "Подрядчик B"),
            (7, "Подрядчик C"),
        ]

        for col, value in cell_data:
            ws.cell(row=6, column=col, value=value)

        result = read_contractors(ws)
        assert result is not None

        # Проверяем, что порядок column_start возрастающий
        column_positions = [cell_info["column_start"] for cell_info in result]
        assert column_positions == sorted(column_positions), "Ячейки должны быть упорядочены по позиции колонки"

        # Проверяем соответствие значений ожидаемому порядку
        expected_values = ["наименование контрагента", "Подрядчик A", "Подрядчик B", "Подрядчик C"]
        actual_values = [cell_info["value"] for cell_info in result]
        assert actual_values == expected_values

    def test_coordinate_format_consistency(self, worksheet_with_contractors):
        """
        ПОВЕДЕНИЕ: Все координаты должны иметь консистентный формат Excel (например, "A1").
        """
        result = read_contractors(worksheet_with_contractors)

        import re

        coordinate_pattern = re.compile(r"^[A-Z]+\d+$")

        for cell_info in result:
            coordinate = cell_info["coordinate"]
            assert coordinate_pattern.match(
                coordinate
            ), f"Координата '{coordinate}' должна соответствовать формату Excel"

    def test_returns_independent_data_structures(self, worksheet_with_contractors):
        """
        ПОВЕДЕНИЕ: Каждый вызов функции должен возвращать
        независимые структуры данных.
        """
        result1 = read_contractors(worksheet_with_contractors)
        result2 = read_contractors(worksheet_with_contractors)

        # Результаты должны быть равны, но не быть одним объектом
        assert result1 == result2
        assert result1 is not result2

        # Изменение одного не должно влиять на другой
        if result1:
            result1[0]["test_field"] = "modified"
            assert "test_field" not in result2[0]
