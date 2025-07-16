"""
Модуль для извлечения информации об исполнителе документа из Excel.

Функция этого модуля сканирует несколько предопределенных строк в нижней
части листа Excel для поиска и извлечения имени исполнителя, его контактного
телефона и даты составления документа. Поиск основан на ключевых фразах,
определенных в константах.
"""

from typing import Dict, Optional

from openpyxl.worksheet.worksheet import Worksheet

# Импорт констант, используемых для ключей JSON и поиска текстовых маркеров
from ..constants import (
    JSON_KEY_EXECUTOR_DATE,
    JSON_KEY_EXECUTOR_NAME,
    JSON_KEY_EXECUTOR_PHONE,
    TABLE_PARSE_EXECUTOR,
    TABLE_PARSE_PREPARATION_DATE,
    TABLE_PARSE_TELEPHONE,
)


def read_executer_block(ws: Worksheet) -> Dict[str, Optional[str]]:
    """
    Извлекает информацию об исполнителе (имя, телефон, дата составления)
    из блока данных, обычно расположенного в нижней части листа Excel.

    Функция сканирует три конкретные строки: `ws.max_row - 5`, `ws.max_row - 4`
    и `ws.max_row - 3`. В каждой из этих строк она проверяет значение во второй
    колонке (колонка 'B') на наличие ключевых фраз (определенных в константах
    `TABLE_PARSE_EXECUTOR`, `TABLE_PARSE_TELEPHONE`, `TABLE_PARSE_PREPARATION_DATE`),
    используя регистронезависимое сравнение.

    -   Для "исполнителя" (`TABLE_PARSE_EXECUTOR`) и "телефона" (`TABLE_PARSE_TELEPHONE`):
        предполагается, что данные следуют за первым двоеточием в строке.
    -   Для "даты составления" (`TABLE_PARSE_PREPARATION_DATE`): предполагается,
        что данные следуют непосредственно за самой ключевой фразой.

    Извлеченные значения очищаются от начальных/конечных пробелов.

    Args:
        ws (Worksheet): Лист Excel (объект openpyxl.worksheet.worksheet.Worksheet),
            из которого считываются данные.

    Returns:
        Dict[str, Optional[str]]: Словарь с тремя ключами:
            - `JSON_KEY_EXECUTOR_NAME`
            - `JSON_KEY_EXECUTOR_PHONE`
            - `JSON_KEY_EXECUTOR_DATE`
            Значениями являются извлеченные строки данных или `None`, если
            соответствующая информация не была найдена или извлечена.

    Пример возвращаемого значения:
        {
            "executor_name": "Иванов И.И.",
            "executor_phone": "+7 (123) 456-78-90",
            "executor_date": "16.05.2025"  // или ": 16.05.2025" в зависимости от форматирования и логики split
        }

    Примечания по реализации:
        - Код корректно обрабатывает случаи, когда значение в ячейке не является строкой
          (используя `isinstance(cell_value_raw, str)`).
        - Код использует `try-except IndexError` для безопасного извлечения данных
          после `.split()`, на случай если ожидаемый разделитель (например, ":") отсутствует.
    """
    max_sheet_row = ws.max_row
    executor_info: Dict[str, Optional[str]] = {
        JSON_KEY_EXECUTOR_NAME: None,
        JSON_KEY_EXECUTOR_PHONE: None,
        JSON_KEY_EXECUTOR_DATE: None,
    }

    # Определяем диапазон строк для сканирования (предпоследние строки)
    # Например, если max_sheet_row = 20, сканируются строки 15, 16, 17.
    # range(max_sheet_row - 5, max_sheet_row - 2) соответствует строкам max_row-5, max_row-4, max_row-3.
    for row_to_scan in range(max_sheet_row - 5, max_sheet_row - 2):
        # Проверка, чтобы индекс строки был валидным (больше 0)
        if row_to_scan < 1:
            continue

        # Значение извлекается из второй колонки (B) текущей сканируемой строки
        cell_value_raw = ws.cell(row=row_to_scan, column=2).value

        if isinstance(cell_value_raw, str):
            # Приводим значение ячейки и константы к нижнему регистру для регистронезависимого сравнения
            cell_value_lower = cell_value_raw.lower()

            executor_prefix_lower = TABLE_PARSE_EXECUTOR.lower()
            phone_prefix_lower = TABLE_PARSE_TELEPHONE.lower()
            date_prefix_lower = TABLE_PARSE_PREPARATION_DATE.lower()

            if cell_value_lower.startswith(executor_prefix_lower):
                try:
                    # Извлекаем текст после первого двоеточия из ОРИГИНАЛЬНОЙ строки
                    executor_info[JSON_KEY_EXECUTOR_NAME] = cell_value_raw.split(
                        ":", 1
                    )[1].strip()
                except IndexError:
                    # Двоеточие не найдено, значение останется None (или предыдущим, если уже было найдено)
                    pass
            elif cell_value_lower.startswith(phone_prefix_lower):
                try:
                    # Извлекаем текст после первого двоеточия из ОРИГИНАЛЬНОЙ строки
                    executor_info[JSON_KEY_EXECUTOR_PHONE] = cell_value_raw.split(
                        ":", 1
                    )[1].strip()
                except IndexError:
                    pass
            elif cell_value_lower.startswith(date_prefix_lower):
                try:
                    # 1. Берём всё, что находится ПОСЛЕ ключевой фразы
                    prefix_len = len(TABLE_PARSE_PREPARATION_DATE)
                    value_part = cell_value_raw[prefix_len:]

                    # 2. Убираем пробелы в начале и смотрим, есть ли теперь впереди двоеточие
                    if value_part.lstrip().startswith(":"):
                        # Если да - это формат "Ключ: Значение".
                        # Находим первое двоеточие и берем всё, что после него.
                        colon_index = value_part.find(":")
                        final_value = value_part[colon_index + 1 :]
                    else:
                        # Если нет - это формат "Ключ Значение".
                        # Значит, value_part уже содержит то, что нам нужно.
                        final_value = value_part

                    # 3. Записываем результат, очищенный от лишних пробелов
                    executor_info[JSON_KEY_EXECUTOR_DATE] = final_value.strip()
                except Exception:
                    # Безопасная обработка на случай любых ошибок
                    pass

    return executor_info
