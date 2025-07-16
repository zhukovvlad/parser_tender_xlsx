"""
Модуль для извлечения детализированных позиций и итоговых сумм подрядчика.

Функция этого модуля обрабатывает строки на листе Excel, относящиеся к одному
подрядчику, разделяя их на обычные товарные/рабочие позиции и итоговые
(суммирующие) строки. Логика переключается в режим сбора итогов при
обнаружении объединенной ячейки в начале строки.
"""

from typing import Dict, Any
from openpyxl.worksheet.worksheet import Worksheet

from .sanitize_text import normalize_job_title_with_lemmatization

# Локальные импорты (из той же директории helpers)
from .get_items_dict import get_items_dict
from .parse_contractor_row import parse_contractor_row

# Импорт констант
from ..constants import (
    START_INDEXING_POSITION_ROW,
    JSON_KEY_NUMBER,
    JSON_KEY_CHAPTER_NUMBER,
    JSON_KEY_ARTICLE_SMR,
    JSON_KEY_JOB_TITLE,
    JSON_KEY_JOB_TITLE_NORMALIZED,
    JSON_KEY_COMMENT_ORGANIZER,
    JSON_KEY_UNIT,
    JSON_KEY_QUANTITY,
    JSON_KEY_CONTRACTOR_POSITIONS,
    JSON_KEY_CONTRACTOR_SUMMARY,
    JSON_KEY_TOTAL_COST_VAT,
    JSON_KEY_VAT,
    JSON_KEY_DEVIATION_FROM_CALCULATED_COST, # Используется и для ключа в summary, и для поиска по тексту
    JSON_KEY_INITIAL_COST,                  # Используется и для ключа в summary, и для поиска по тексту
    TABLE_PARSE_DEVIATION_FROM_CALCULATED_COST, # Текстовый маркер для поиска "отклонение..."
    TABLE_PARSE_INITIAL_COST                   # Текстовый маркер для поиска "первоначальная..."
)


def get_positions(ws: Worksheet, contractor: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Извлекает все детализированные позиции работ/материалов и отдельно
    итоговые/суммирующие строки, относящиеся к конкретному подрядчику,
    из листа Excel.

    Функция начинает чтение данных с предопределенной строки
    (константа `START_INDEXING_POSITION_ROW`) и продолжает до тех пор,
    пока не встретит полностью пустую строку или не выйдет за пределы листа.

    Различает два режима обработки строк:
    1.  'normal': Для обычных строк с детализированными позициями.
        Эти позиции добавляются в словарь, возвращаемый по ключу
        `JSON_KEY_CONTRACTOR_POSITIONS`.
    2.  'merged_mode': Для итоговых/суммирующих строк (например, "ИТОГО", "НДС"),
        которые часто начинаются с объединенной ячейки. Эти строки добавляются
        в словарь, возвращаемый по ключу `JSON_KEY_CONTRACTOR_SUMMARY`.

    Переключение в 'merged_mode' происходит, если первая ячейка (колонка А)
    текущей строки является частью объединенного диапазона, и функция ранее
    была в режиме 'normal'. После переключения функция остается в 'merged_mode'.

    Args:
        ws (Worksheet): Лист Excel (объект openpyxl.worksheet.worksheet.Worksheet),
            с которого считываются данные.
        contractor (Dict[str, Any]): Словарь, содержащий информацию о подрядчике.
            Должен включать как минимум:
            - "merged_shape" (Dict[str, int]): Содержит {"colspan": int}, указывающее
              количество колонок, занимаемых данными этого подрядчика.
            - "column_start" (int): Номер начальной колонки данных этого подрядчика.
            Опционально может содержать "value" (str) для отладочного вывода имени подрядчика.

    Returns:
        Dict[str, Dict[str, Any]]: Словарь с двумя основными ключами:
        - `JSON_KEY_CONTRACTOR_POSITIONS` (например, "positions"): Словарь, где
          ключами являются порядковые номера позиций в виде строк ("1", "2", ...),
          а значениями — словари, описывающие каждую позицию. Структура каждого
          такого словаря создается `get_items_dict` и дополняется данными,
          возвращаемыми `parse_contractor_row`.
          Итоговые строки сюда НЕ попадают.
        - `JSON_KEY_CONTRACTOR_SUMMARY` (например, "summary"): Словарь, где
          ключами являются строки, идентифицирующие итоговые показатели
          (например, `JSON_KEY_TOTAL_COST_VAT`, `JSON_KEY_VAT`,
          `JSON_KEY_DEVIATION_FROM_CALCULATED_COST`, `JSON_KEY_INITIAL_COST`,
          или сгенерированный ключ вида "merged_{номер_строки}" для нераспознанных),
          а значениями — словари, описывающие эти итоговые строки. Каждая
          итоговая строка содержит поле `JSON_KEY_JOB_TITLE` (текст из первой
          ячейки строки) и данные, возвращаемые `parse_contractor_row`.
          Эти строки НЕ дублируются в словаре позиций.

    Side effects:
        - Печатает отладочную информацию в консоль на каждой итерации обработки строки,
          если соответствующий `print` не закомментирован.
    """
    # Отладочный print в начале функции, если нужен:
    # print(f"==== START get_positions for contractor: {contractor.get('value', 'N/A')} ====")

    positions: Dict[str, Any] = {}
    summary: Dict[str, Any] = {}
    current_row_num: int = START_INDEXING_POSITION_ROW
    
    contractor_colspan: int = contractor["merged_shape"]["colspan"]
    
    mode: str = "normal"  # 'normal' или 'merged_mode'
    item_index: int = 1   # Счетчик только для детализированных позиций в режиме 'normal'

    while True:
        if current_row_num > ws.max_row: # Защита от выхода за пределы данных листа
            break

        current_row_tuple = ws[current_row_num] # Кортеж ячеек (Cell) для текущей строки
        if all(cell.value is None for cell in current_row_tuple):
            break # Прерываем цикл, если строка полностью пустая

        first_cell_in_row = current_row_tuple[0] # Ячейка в первой колонке (A)


        print(
            f"get_positions: row={current_row_num}, mode='{mode}', "
            f"first_cell_value='{first_cell_in_row.value}', "
            f"contractor='{contractor.get('value', 'N/A')}'"
        )

        if mode == "normal":
            is_first_cell_merged = False
            for merged_range_obj in ws.merged_cells.ranges:
                if (merged_range_obj.min_row <= first_cell_in_row.row <= merged_range_obj.max_row and
                        merged_range_obj.min_col <= first_cell_in_row.column <= merged_range_obj.max_col):
                    is_first_cell_merged = True
                    break
            if is_first_cell_merged:
                mode = "merged_mode"

        if mode == "normal":
            item = get_items_dict(contractor_colspan)
            item[JSON_KEY_NUMBER] = ws.cell(row=current_row_num, column=1).value
            item[JSON_KEY_CHAPTER_NUMBER] = ws.cell(row=current_row_num, column=2).value
            item[JSON_KEY_ARTICLE_SMR] = ws.cell(row=current_row_num, column=3).value
            original_job_title = ws.cell(row=current_row_num, column=4).value
            item[JSON_KEY_JOB_TITLE] = original_job_title
            item[JSON_KEY_JOB_TITLE_NORMALIZED] = normalize_job_title_with_lemmatization(original_job_title)
            item[JSON_KEY_COMMENT_ORGANIZER] = ws.cell(row=current_row_num, column=6).value
            item[JSON_KEY_UNIT] = ws.cell(row=current_row_num, column=7).value
            item[JSON_KEY_QUANTITY] = ws.cell(row=current_row_num, column=8).value

            contractor_specific_data = parse_contractor_row(ws, current_row_num, contractor)
            item.update(contractor_specific_data)
            
            positions[str(item_index)] = item
            item_index += 1
        
        else: # mode == "merged_mode"
            summary_label_raw = str(first_cell_in_row.value).strip().lower() if first_cell_in_row.value is not None else ""
            summary_key = f"merged_{current_row_num}"

            if "итого" in summary_label_raw and "ндс" in summary_label_raw:
                summary_key = JSON_KEY_TOTAL_COST_VAT
            elif "в том числе ндс" in summary_label_raw or \
                 ("ндс" in summary_label_raw and "итого" not in summary_label_raw):
                summary_key = JSON_KEY_VAT
            elif TABLE_PARSE_DEVIATION_FROM_CALCULATED_COST in summary_label_raw:
                summary_key = JSON_KEY_DEVIATION_FROM_CALCULATED_COST
            elif TABLE_PARSE_INITIAL_COST in summary_label_raw:
                summary_key = JSON_KEY_INITIAL_COST
            
            summary_item_data: Dict[str, Any] = {
                JSON_KEY_JOB_TITLE: first_cell_in_row.value,
            }
            contractor_specific_summary_data = parse_contractor_row(ws, current_row_num, contractor)
            summary_item_data.update(contractor_specific_summary_data)
            
            summary[summary_key] = summary_item_data

        current_row_num += 1

    return {
        JSON_KEY_CONTRACTOR_POSITIONS: positions,
        JSON_KEY_CONTRACTOR_SUMMARY: summary
    }