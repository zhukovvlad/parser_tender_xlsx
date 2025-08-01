# helpers/get_lot_positions.py

"""Модуль для извлечения детализированных позиций в границах одного лота.

Назначение:
    Этот модуль предоставляет специализированную функцию `get_lot_positions` для
    парсинга основной части тендерной таблицы в Excel. Его главная задача —
    считать все строки с работами/материалами (позициями), которые относятся
    к одному подрядчику и находятся строго в границах одного лота.

    Модуль не занимается поиском итоговых (summary) или общих информационных
    блоков. Он спроектирован для вызова из более высокоуровневых функций
    (таких как `get_proposals`), которые предварительно определяют границы
    (начальную и конечную строки) для каждого лота.

Ключевые зависимости:
    - `parse_contractor_row`: для извлечения данных из колонок, специфичных
      для подрядчика (стоимости, предлагаемое количество и т.д.).
    - `normalize_job_title_with_lemmatization`: для очистки и нормализации
      наименования работ.
"""

import logging
from typing import Any, Dict

from openpyxl.worksheet.worksheet import Worksheet

from ..constants import (
    JSON_KEY_ARTICLE_SMR,
    JSON_KEY_CHAPTER_NUMBER,
    JSON_KEY_COMMENT_ORGANIZER,
    JSON_KEY_JOB_TITLE,
    JSON_KEY_JOB_TITLE_NORMALIZED,
    JSON_KEY_NUMBER,
    JSON_KEY_QUANTITY,
    JSON_KEY_UNIT,
    START_INDEXING_POSITION_ROW,
)
from .get_items_dict import get_items_dict
from .parse_contractor_row import parse_contractor_row
from .sanitize_text import normalize_job_title_with_lemmatization

log = logging.getLogger(__name__)


def get_lot_positions(
    ws: Worksheet, contractor: Dict[str, Any], lot_start_row: int, lot_end_row: int
) -> Dict[str, Any]:
    """
    Извлекает детализированные позиции для подрядчика строго в границах лота.

    Функция сканирует лист Excel в диапазоне строк, определённом `lot_start_row`
    и `lot_end_row`, и собирает информацию по каждой строке-позиции.

    Логика работы:
    1.  Определяет начальную строку для сканирования как максимальное значение
        из `START_INDEXING_POSITION_ROW` (глобальное начало таблицы) и
        `lot_start_row` (начало текущего лота).
    2.  Итерируется по строкам до `lot_end_row`.
    3.  **Досрочно прекращает сканирование**, если встречает объединённую ячейку
        в первой колонке. Это является надёжным признаком окончания блока
        детализированных позиций и начала блока общих итогов (`summary`).
    4.  Для каждой валидной строки-позиции:
        - Считывает общие данные из фиксированных колонок (номер, наименование,
          единицы измерения, количество и т.д.).
        - Вызывает `parse_contractor_row` для считывания данных из колонок,
          специфичных для данного подрядчика (стоимости и пр.).
        - Объединяет все данные в единый словарь для этой позиции.
    5.  Собирает все словари позиций в один общий словарь, где ключами
        являются порядковые номера "1", "2", "3" и т.д.

    Args:
        ws (Worksheet): Активный лист Excel для анализа.
        contractor (Dict[str, Any]): Словарь с информацией о подрядчике,
            должен содержать ключи "merged_shape" и "column_start".
        lot_start_row (int): Номер начальной строки лота на листе.
        lot_end_row (int): Номер конечной строки лота на листе.

    Returns:
        Dict[str, Any]: Словарь, где ключи - это строковые порядковые номера
        позиций, а значения - словари с полной информацией по каждой позиции.
        Пример:
        {
            "1": {"number": "1.1", "job_title": "Работа 1", ...},
            "2": {"number": "1.2", "job_title": "Работа 2", ...}
        }
    """
    positions: Dict[str, Any] = {}
    item_index: int = 1

    start_scan_row = max(START_INDEXING_POSITION_ROW, lot_start_row)

    for current_row_num in range(start_scan_row, lot_end_row + 1):
        log.debug(
            f"get_lot_positions: Обработка строки {current_row_num} "
            f"для подрядчика '{contractor.get('value', 'N/A')}' "
            f"в границах лота [{lot_start_row}-{lot_end_row}]"
        )
        # Проверяем, не является ли первая ячейка объединенной,
        # что является признаком начала блока итогов (summary).
        first_cell = ws.cell(row=current_row_num, column=1)
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if first_cell.coordinate in merged_range:
                is_merged = True
                break

        # Если мы наткнулись на объединенную ячейку, значит, блок позиций закончился.
        if is_merged:
            break

        # Пропускаем полностью пустые строки
        current_row_tuple = ws[current_row_num]
        if all(cell.value is None for cell in current_row_tuple):
            continue

        item = get_items_dict(contractor["merged_shape"]["colspan"])
        item[JSON_KEY_NUMBER] = ws.cell(row=current_row_num, column=1).value
        item[JSON_KEY_CHAPTER_NUMBER] = ws.cell(row=current_row_num, column=2).value
        item[JSON_KEY_ARTICLE_SMR] = ws.cell(row=current_row_num, column=3).value
        original_job_title = ws.cell(row=current_row_num, column=4).value
        item[JSON_KEY_JOB_TITLE] = original_job_title
        item[JSON_KEY_JOB_TITLE_NORMALIZED] = normalize_job_title_with_lemmatization(original_job_title)
        item[JSON_KEY_COMMENT_ORGANIZER] = ws.cell(row=current_row_num, column=6).value
        item[JSON_KEY_UNIT] = ws.cell(row=current_row_num, column=7).value
        item[JSON_KEY_QUANTITY] = ws.cell(row=current_row_num, column=8).value

        contractor_specific_data = parse_contractor_row(ws, current_row_num, contractor)
        item.update(contractor_specific_data)

        positions[str(item_index)] = item
        item_index += 1

    return positions
