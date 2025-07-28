# helpers/get_proposals.py

"""Модуль-оркестратор для агрегации предложений подрядчиков.

Назначение:
    Этот модуль выполняет роль "дирижера". Его основная функция `get_proposals`
    отвечает за сборку полной информации о предложениях всех подрядчиков,
    но делает это для одного конкретного лота, границы которого (`start_row`,
    `end_row`) ей передаются.

    Для этого она вызывает ряд более специализированных функций:
    - `read_contractors`: для получения общего списка всех подрядчиков
      из заголовка таблицы.
    - `get_lot_positions`: для сбора детализированных позиций, которые
      находятся строго в границах переданного лота.
    - `get_summary`: для сбора общих итоговых сумм, которые находятся
      в конце таблицы и относятся ко всему тендеру.
    - `get_additional_info`: для сбора дополнительной информации, также
      относящейся ко всему тендеру.
"""

from typing import Dict, Any, List, Optional
from openpyxl.worksheet.worksheet import Worksheet

# --- ИЗМЕНЕНИЕ 1: Обновлены импорты ---
from .get_additional_info import get_additional_info
from .get_lot_positions import get_lot_positions
from .get_summary import get_summary
from .read_contractors import read_contractors

# Импорт констант
from ..constants import (
    JSON_KEY_CONTRACTOR_ACCREDITATION,
    JSON_KEY_CONTRACTOR_ADDITIONAL_INFO,
    JSON_KEY_CONTRACTOR_ADDRESS,
    JSON_KEY_CONTRACTOR_COORDINATE,
    JSON_KEY_CONTRACTOR_HEIGHT,
    JSON_KEY_CONTRACTOR_INDEX,
    JSON_KEY_CONTRACTOR_INN,
    JSON_KEY_CONTRACTOR_ITEMS,
    JSON_KEY_CONTRACTOR_TITLE,
    JSON_KEY_CONTRACTOR_WIDTH,
    JSON_KEY_CONTRACTOR_POSITIONS,
    JSON_KEY_CONTRACTOR_SUMMARY
)

def get_proposals(ws: Worksheet, start_row: int, end_row: int) -> Dict[str, Dict[str, Any]]:
    """Собирает предложения всех подрядчиков для одного конкретного лота.

    Функция выступает в роли агрегатора, который для каждого подрядчика,
    найденного на листе, собирает полный набор данных, корректно разделяя
    информацию, привязанную к лоту, и информацию, общую для всего тендера.

    Логика работы:
    1.  С помощью `read_contractors(ws)` получает полный список всех
        подрядчиков, присутствующих в документе.
    2.  Итерируется по этому списку.
    3.  Для каждого подрядчика:
        а. Извлекает базовую информацию (ИНН, адрес и т.д.) из строк,
           следующих сразу за заголовком подрядчика.
        б. Вызывает `get_lot_positions()`, передавая ей границы `start_row`
           и `end_row`, чтобы получить список позиций, относящихся
           **только к текущему лоту**.
        в. Вызывает `get_summary()`, которая сканирует нижнюю часть таблицы
           и возвращает общие итоговые суммы **по всему тендеру**.
        г. Вызывает `get_additional_info()` для получения дополнительной
           информации, также общей для всего тендера.
        д. Собирает все полученные данные в единый словарь-предложение
           для данного подрядчика.

    Args:
        ws (Worksheet): Активный лист Excel для анализа.
        start_row (int): Номер начальной строки лота на листе.
        end_row (int): Номер конечной строки лота на листе.

    Returns:
        Dict[str, Dict[str, Any]]: Словарь, где ключи - это идентификаторы
        подрядчиков ("contractor_1", "contractor_2", ...), а значения -
        словари с полной информацией по их предложениям.
        Важно: `positions` в результате будут уникальны для каждого лота,
        а `summary` будет одинаковым.
    """
    contractors_list: Optional[List[Dict[str, Any]]] = read_contractors(ws)
    proposals: Dict[str, Dict[str, Any]] = {}

    if not contractors_list:
        return proposals

    for i in range(1, len(contractors_list)):
        contractor_details = contractors_list[i]

        contractor_name: Optional[str] = contractor_details.get("value")
        contractor_row_start: Optional[int] = contractor_details.get("row_start")
        contractor_col_start: Optional[int] = contractor_details.get("column_start")
        contractor_coordinate: Optional[str] = contractor_details.get("coordinate")

        merged_shape: Dict[str, int] = contractor_details.get("merged_shape", {})
        rowspan: int = merged_shape.get("rowspan", 1)
        colspan: int = merged_shape.get("colspan", 1)

        inn_val: Optional[str] = None
        address_val: Optional[str] = None
        accreditation_val: Optional[str] = None

        if rowspan == 1 and contractor_row_start is not None and contractor_col_start is not None:
            inn_val = ws.cell(row=contractor_row_start + 1, column=contractor_col_start).value
            address_val = ws.cell(row=contractor_row_start + 2, column=contractor_col_start).value
            accreditation_val = ws.cell(row=contractor_row_start + 3, column=contractor_col_start).value
        
        # --- ИЗМЕНЕНИЕ 2: Раздельные вызовы для позиций и итогов ---
        
        # 1. Получаем позиции, относящиеся только к текущему лоту
        positions_data = get_lot_positions(
            ws,
            contractor_details,
            lot_start_row=start_row,
            lot_end_row=end_row
        )
        
        # 2. Получаем общие итоги по всему тендеру
        summary_data = get_summary(ws, contractor_details)
        
        # 3. Собираем их в единую структуру
        contractor_items_data = {
            JSON_KEY_CONTRACTOR_POSITIONS: positions_data,
            JSON_KEY_CONTRACTOR_SUMMARY: summary_data
        }
        
        # 4. Получаем дополнительную информацию (не зависит от лотов)
        contractor_additional_info_data = get_additional_info(ws, contractor_details)
        
        proposal_key = f"{JSON_KEY_CONTRACTOR_INDEX}{i}"
        proposals[proposal_key] = {
            JSON_KEY_CONTRACTOR_TITLE: contractor_name,
            JSON_KEY_CONTRACTOR_INN: inn_val,
            JSON_KEY_CONTRACTOR_ADDRESS: address_val,
            JSON_KEY_CONTRACTOR_ACCREDITATION: accreditation_val,
            JSON_KEY_CONTRACTOR_COORDINATE: contractor_coordinate,
            JSON_KEY_CONTRACTOR_WIDTH: colspan,
            JSON_KEY_CONTRACTOR_HEIGHT: rowspan,
            JSON_KEY_CONTRACTOR_ITEMS: contractor_items_data,
            JSON_KEY_CONTRACTOR_ADDITIONAL_INFO: contractor_additional_info_data,
        }
            
    return proposals