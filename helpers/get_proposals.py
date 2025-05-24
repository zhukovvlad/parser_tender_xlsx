"""
Модуль для агрегации предложений всех подрядчиков с листа Excel.

Этот модуль использует другие вспомогательные функции для сбора полной
информации по каждому подрядчику, включая его основные данные,
детализированные позиции по работам/материалам, и дополнительную информацию.
Результатом является структурированный словарь всех предложений.
"""

from typing import Dict, Any, List, Optional
from openpyxl.worksheet.worksheet import Worksheet

# Локальные импорты (из той же директории helpers)
from .get_additional_info import get_additional_info
from .get_positions import get_positions
from .read_contractors import read_contractors # Эта функция возвращает список словарей подрядчиков

# Импорт констант для ключей JSON
from constants import (
    JSON_KEY_CONTRACTOR_ACCREDITATION,
    JSON_KEY_CONTRACTOR_ADDITIONAL_INFO,
    JSON_KEY_CONTRACTOR_ADDRESS,
    JSON_KEY_CONTRACTOR_COORDINATE,
    JSON_KEY_CONTRACTOR_HEIGHT,      # Используется для 'rowspan'
    JSON_KEY_CONTRACTOR_INDEX,
    JSON_KEY_CONTRACTOR_INN,
    JSON_KEY_CONTRACTOR_ITEMS,
    JSON_KEY_CONTRACTOR_TITLE,
    JSON_KEY_CONTRACTOR_WIDTH        # Используется для 'colspan'
)

def get_proposals(ws: Worksheet) -> Dict[str, Dict[str, Any]]:
    """
    Извлекает и структурирует данные по предложениям всех подрядчиков с листа Excel.

    Сначала с помощью `read_contractors(ws)` получается список словарей,
    каждый из которых описывает заголовок одного подрядчика. Предполагается,
    что первый элемент этого списка (`contractors_list[0]`) может быть общим
    заголовком колонки (например, "Наименование контрагента") и пропускается.

    Для каждого последующего подрядчика (начиная с `contractors_list[1]`):
    1.  Извлекаются основные данные: имя (`value`), координаты (`coordinate`),
        информация об объединении ячеек (`merged_shape`, если есть) из словаря подрядчика.
    2.  На основе `merged_shape` определяются `rowspan` (высота) и `colspan` (ширина).
    3.  Если `rowspan == 1`, извлекаются ИНН, адрес и сведения об аккредитации
        из ячеек, расположенных со смещением +1, +2, +3 строки относительно
        `contractor["row_start"]` и в той же колонке `contractor["column_start"]`.
        В противном случае, эти поля устанавливаются в `None`.
    4.  Вызываются `get_positions(ws, contractor_details)` для получения
        детализированных позиций и итогов.
    5.  Вызываются `get_additional_info(ws, contractor_details)` для получения
        дополнительной информации.

    Результат собирается в словарь, где ключи — это идентификаторы подрядчиков
    (например, "contractor_1"), а значения — словари с полной информацией.

    Args:
        ws (Worksheet): Лист Excel (объект openpyxl.worksheet.worksheet.Worksheet),
            с которого считываются данные.

    Returns:
        Dict[str, Dict[str, Any]]: Словарь, где ключи - это строковые
            идентификаторы подрядчиков (например, "contractor_1", "contractor_2", ...),
            а значения - словари с подробной информацией по каждому подрядчику.
            Структура информации для каждого подрядчика:
            {
                JSON_KEY_CONTRACTOR_TITLE (str): Наименование подрядчика,
                JSON_KEY_CONTRACTOR_INN (Optional[str]): ИНН,
                JSON_KEY_CONTRACTOR_ADDRESS (Optional[str]): Адрес,
                JSON_KEY_CONTRACTOR_ACCREDITATION (Optional[str]): Сведения об аккредитации,
                JSON_KEY_CONTRACTOR_COORDINATE (Optional[str]): Координата начальной ячейки заголовка подрядчика,
                JSON_KEY_CONTRACTOR_WIDTH (int): Ширина объединенной ячейки (colspan),
                JSON_KEY_CONTRACTOR_HEIGHT (int): Высота объединенной ячейки (rowspan),
                JSON_KEY_CONTRACTOR_ITEMS (Dict): Структура, возвращаемая get_positions(),
                JSON_KEY_CONTRACTOR_ADDITIONAL_INFO (Dict): Структура, возвращаемая get_additional_info()
            }
            Если подрядчики не найдены, возвращается пустой словарь.
    """
    # Получаем список словарей, каждый из которых представляет заголовок подрядчика
    # (или общий заголовок "Наименование контрагента" как первый элемент).
    contractors_list: Optional[List[Dict[str, Any]]] = read_contractors(ws)
    proposals: Dict[str, Dict[str, Any]] = {}

    if not contractors_list:
        return proposals # Нет подрядчиков - нет предложений

    # Итерация по списку подрядчиков, начиная со второго элемента (индекс 1),
    # так как первый элемент (индекс 0) часто является общим заголовком колонки.
    for i in range(1, len(contractors_list)):
        contractor_details = contractors_list[i]

        contractor_name: Optional[str] = contractor_details.get("value")
        contractor_row_start: Optional[int] = contractor_details.get("row_start")
        contractor_col_start: Optional[int] = contractor_details.get("column_start")
        contractor_coordinate: Optional[str] = contractor_details.get("coordinate")

        # Информация об объединении ячейки (если есть)
        merged_shape: Dict[str, int] = contractor_details.get("merged_shape", {})
        rowspan: int = merged_shape.get("rowspan", 1) # Высота ячейки (по умолчанию 1)
        colspan: int = merged_shape.get("colspan", 1) # Ширина ячейки (по умолчанию 1)

        inn_val: Optional[str] = None
        address_val: Optional[str] = None
        accreditation_val: Optional[str] = None

        # ИНН, адрес и аккредитация извлекаются только если заголовок подрядчика не объединен по строкам (rowspan == 1)
        # и если известны начальная строка и колонка.
        if rowspan == 1 and contractor_row_start is not None and contractor_col_start is not None:
            inn_val = ws.cell(row=contractor_row_start + 1, column=contractor_col_start).value
            address_val = ws.cell(row=contractor_row_start + 2, column=contractor_col_start).value
            accreditation_val = ws.cell(row=contractor_row_start + 3, column=contractor_col_start).value
        
        # Получение детализированных позиций и дополнительной информации для текущего подрядчика
        # Функции get_positions и get_additional_info ожидают словарь contractor_details
        # с ключами "column_start" и "merged_shape".
        contractor_items_data = get_positions(ws, contractor_details)
        contractor_additional_info_data = get_additional_info(ws, contractor_details)
        
        # Формирование итогового словаря для данного подрядчика
        # Ключ генерируется как "contractor_1", "contractor_2", и т.д. (i начинается с 1)
        proposal_key = f"{JSON_KEY_CONTRACTOR_INDEX}{i}"
        proposals[proposal_key] = {
            JSON_KEY_CONTRACTOR_TITLE: contractor_name,
            JSON_KEY_CONTRACTOR_INN: inn_val,
            JSON_KEY_CONTRACTOR_ADDRESS: address_val,
            JSON_KEY_CONTRACTOR_ACCREDITATION: accreditation_val,
            JSON_KEY_CONTRACTOR_COORDINATE: contractor_coordinate,
            JSON_KEY_CONTRACTOR_WIDTH: colspan,    # Ширина (colspan)
            JSON_KEY_CONTRACTOR_HEIGHT: rowspan,   # Высота (rowspan)
            JSON_KEY_CONTRACTOR_ITEMS: contractor_items_data,
            JSON_KEY_CONTRACTOR_ADDITIONAL_INFO: contractor_additional_info_data,
        }
            
    return proposals