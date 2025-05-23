"""
Модуль для работы с объединенными ячейками в листах Excel.

Этот модуль предоставляет функцию для построения карты объединенных ячеек,
которая позволяет легко определять размеры (rowspan, colspan) объединенного
диапазона для любой ячейки, входящей в этот диапазон.
Это полезно при парсинге сложных Excel-таблиц, где структура ячеек
имеет значение.
"""
from openpyxl.worksheet.worksheet import Worksheet


def build_merged_shape_map(ws: Worksheet) -> dict:
    """
    Создает и возвращает карту размеров объединенных ячеек на листе Excel.

    Эта функция проходит по всем диапазонам объединенных ячеек на предоставленном
    листе `openpyxl.worksheet.worksheet.Worksheet`. Для каждого такого диапазона
    она вычисляет количество объединенных строк (rowspan) и столбцов (colspan).
    Затем для каждой отдельной ячейки, входящей в этот объединенный диапазон,
    в итоговый словарь добавляется запись, где ключом является координата ячейки
    (например, "A1", "B2"), а значением — словарь с информацией о rowspan и colspan
    всего объединенного диапазона, к которому эта ячейка принадлежит.

    Это полезно для определения фактических размеров объединенной ячейки,
    когда известна координата любой из входящих в нее ячеек.

    Args:
        ws (Worksheet): Активный лист Excel (объект openpyxl.worksheet.worksheet.Worksheet),
        из которого будут считываться данные об объединенных ячейках.

    Returns:
        dict: Словарь, где:
            - ключ (str): Координата ячейки (например, "A1"), которая является частью
                          объединенного диапазона.
            - значение (dict): Словарь с двумя ключами:
                - "rowspan" (int): Количество строк, которые занимает объединенная
                                   ячейка (включительно).
                - "colspan" (int): Количество столбцов, которые занимает объединенная
                                   ячейка (включительно).
    
    Пример возвращаемого значения для листа, где ячейки A1:C2 объединены:
    {
        "A1": {"rowspan": 2, "colspan": 3},
        "B1": {"rowspan": 2, "colspan": 3},
        "C1": {"rowspan": 2, "colspan": 3},
        "A2": {"rowspan": 2, "colspan": 3},
        "B2": {"rowspan": 2, "colspan": 3},
        "C2": {"rowspan": 2, "colspan": 3}
        # ... и так далее для других объединенных диапазонов на листе
    }
    """
    merged_map = {}
    # ws.merged_cells.ranges возвращает список объектов MergedCellRange,
    # каждый из которых описывает один объединенный диапазон.
    if hasattr(ws, 'merged_cells') and hasattr(ws.merged_cells, 'ranges'):
        for merged_range in ws.merged_cells.ranges:
            # Вычисляем rowspan и colspan для текущего объединенного диапазона.
            # Координаты в merged_range являются 1-индексированными.
            rowspan = merged_range.max_row - merged_range.min_row + 1
            colspan = merged_range.max_col - merged_range.min_col + 1
            
            # Итерируемся по всем ячейкам внутри этого объединенного диапазона.
            # ws[merged_range.coord] возвращает кортеж кортежей ячеек (Cell objects),
            # представляющих строки и ячейки в указанном диапазоне.
            for row_of_cells in ws[merged_range.coord]:
                for cell in row_of_cells:
                    # Каждой ячейке в объединенном диапазоне сопоставляем
                    # вычисленные размеры всего этого диапазона.
                    merged_map[cell.coordinate] = {"rowspan": rowspan, "colspan": colspan}
    return merged_map